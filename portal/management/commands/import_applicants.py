from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from portal.models import Applicant, ImportBatch


class Command(BaseCommand):
    help = "Import applicants excel (A..I)"

    def add_arguments(self, parser):
        parser.add_argument("path", type=str)

    def handle(self, *args, **opts):
        path = opts["path"]
        wb = load_workbook(path)
        ws = wb.active

        batch = ImportBatch.objects.create(kind="applicants", file_name=path)

        created = updated = 0

        # نفترض الصف الأول عناوين
        for row in ws.iter_rows(min_row=2, values_only=True):
            # A..I
            full_name = (row[0] or "").strip()
            national_id = str(row[1] or "").strip()
            mobile = str(row[2] or "").strip()
            gender = (row[3] or "").strip()
            current_job = (row[4] or "").strip()
            sector = (row[5] or "").strip()
            rank = (row[6] or "").strip()
            start_date = str(row[7] or "").strip()
            current_school = (row[8] or "").strip()

            if not national_id:
                continue

            _, was_created = Applicant.objects.update_or_create(
                national_id=national_id,
                defaults=dict(
                    full_name=full_name,
                    mobile=mobile,
                    gender=gender,
                    current_job=current_job,
                    sector=sector,
                    rank=rank,
                    start_date=start_date,
                    current_school=current_school,
                    batch=batch,
                    is_active=True,
                ),
            )
            created += 1 if was_created else 0
            updated += 0 if was_created else 1

        self.stdout.write(self.style.SUCCESS(f"Applicants imported: created={created}, updated={updated}, batch={batch.id}"))