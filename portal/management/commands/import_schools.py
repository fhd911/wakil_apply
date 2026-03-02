from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from portal.models import SchoolVacancy, ImportBatch


def to_int(v):
    try:
        return int(v or 0)
    except Exception:
        return 0


class Command(BaseCommand):
    help = "Import schools/vacancies excel (A..R)"

    def add_arguments(self, parser):
        parser.add_argument("path", type=str)

    def handle(self, *args, **opts):
        path = opts["path"]
        wb = load_workbook(path)
        ws = wb.active

        batch = ImportBatch.objects.create(kind="schools", file_name=path)

        created = updated = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            # A..R (0..17)
            ministry_no = str(row[0] or "").strip()
            school_name = (row[1] or "").strip()
            stage = (row[2] or "").strip()
            sector = (row[3] or "").strip()
            establishment_status = (row[4] or "").strip()
            gender = (row[5] or "").strip()
            education_type = (row[6] or "").strip()
            manager_national_id = str(row[7] or "").strip()
            manager_name = (row[8] or "").strip()

            students_total = to_int(row[9])
            classes_total = to_int(row[10])
            students_metric = to_int(row[11])
            class_metric = to_int(row[12])
            stage_code = str(row[13] or "").strip()
            stage_metric = to_int(row[14])

            deputy_staff = to_int(row[15])
            deputy_existing = to_int(row[16])
            deputy_need = to_int(row[17])

            if not school_name:
                continue

            # مفتاح التحديث: الأفضل الرقم الوزاري إذا موجود
            key = ministry_no or school_name

            _, was_created = SchoolVacancy.objects.update_or_create(
                ministry_no=key,
                defaults=dict(
                    school_name=school_name,
                    stage=stage,
                    sector=sector,
                    establishment_status=establishment_status,
                    gender=gender,
                    education_type=education_type,
                    manager_national_id=manager_national_id,
                    manager_name=manager_name,
                    students_total=students_total,
                    classes_total=classes_total,
                    students_metric=students_metric,
                    class_metric=class_metric,
                    stage_code=stage_code,
                    stage_metric=stage_metric,
                    deputy_staff=deputy_staff,
                    deputy_existing=deputy_existing,
                    deputy_need=deputy_need,
                    is_open=True,
                    batch=batch,
                ),
            )
            created += 1 if was_created else 0
            updated += 0 if was_created else 1

        self.stdout.write(self.style.SUCCESS(f"Schools imported: created={created}, updated={updated}, batch={batch.id}"))