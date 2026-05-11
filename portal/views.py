from __future__ import annotations

import csv
import os
from datetime import datetime, timedelta
from io import BytesIO
from urllib.parse import urlencode, quote

from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import (
    Q,
    Count,
    OuterRef,
    Subquery,
    Value,
    CharField,
    Case,
    When,
    IntegerField,
)
from django.db.models.functions import Coalesce, Concat, Cast
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.views.decorators.http import require_http_methods, require_POST, require_GET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .forms import ImportExcelForm
from .forms_admin import ApplicantAdminForm, VacancyAdminForm
from .models import (
    Applicant,
    SchoolVacancy,
    Application,
    ApplicationPreference,
    PortalWindow,
    ApplicantDataIssue,
)
from .services_import import import_applicants_xlsx, import_schools_xlsx


SESSION_KEY = "applicant_nid"

# =========================================================
# Submission Policy / نصوص الإقرارات المعتمدة
# =========================================================
SUBMISSION_POLICY_VERSION = "v1"

PREFERENCES_ACK_TEXT = (
    "أقرّ بأن اختياري وترتيبي للرغبات لا يعني استحقاق التوجيه عليها أو تحققها، "
    "ولا يترتب عليه أي التزام بتوجيهي إلى أي منها في حال وجود مرشحين أعلى درجة "
    "أو أحق في المفاضلة، وأن التوجيه النهائي يكون وفق المصلحة التعليمية واحتياج الإدارة "
    "والضوابط المعتمدة ونتائج المفاضلة، وفي حدود الرغبات المحددة."
)

NO_PREFERENCES_ACK_TEXT = (
    "أقرّ بأنني اطلعت على الشواغر المتاحة خلال فترة التقديم، وأرغب في إرسال طلبي "
    "دون اختيار أي رغبة، وأعلم أن الطلب في هذه الحالة يُعد مستلمًا دون رغبات، "
    "ولا يدخل في مفاضلة الرغبات، ولا يترتب عليه أي مطالبة بشاغر محدد، "
    "مع احتفاظ الإدارة بحق معالجة الطلب وفق المصلحة التعليمية والاحتياج والضوابط المعتمدة."
)

NO_PREFERENCES_POLICY_MEANING = (
    "مستلم دون رغبات؛ لا يدخل في مفاضلة الرغبات، ولا يترتب عليه مطالبة بشاغر محدد، "
    "مع احتفاظ الإدارة بحق معالجة الطلب وفق المصلحة التعليمية والاحتياج والضوابط المعتمدة."
)

NO_PREFERENCES_ADMIN_DECISION_NOTE = (
    "تم اعتماد استلام الطلب دون رغبات، ولا يدخل في مفاضلة الرغبات، "
    "ولا يترتب عليه مطالبة بشاغر محدد، مع احتفاظ الإدارة بحق معالجة الطلب "
    "وفق المصلحة التعليمية والاحتياج والضوابط المعتمدة."
)

PREFERENCES_COMPETITION_NOTE = (
    "يدخل مفاضلة الرغبات وفق الضوابط ونتائج المفاضلة والاحتياج، "
    "دون أن يعني ذلك تحقق أي رغبة أو أولوية على الأعلى درجة."
)


# =========================================================
# Helpers
# =========================================================
def _fmt_dt(dt) -> str:
    if not dt:
        return ""
    try:
        return timezone.localtime(dt).strftime("%Y-%m-%d %H:%M")
    except Exception:
        try:
            return dt.strftime("%Y-%m-%d %H:%M")
        except Exception:
            return str(dt)


def _get_applicant(request):
    nid = request.session.get(SESSION_KEY)
    if not nid:
        return None
    return Applicant.objects.filter(national_id=nid, is_active=True).first()


def _portal_gate():
    """
    يرجع:
    (open_now, msg, win)

    open_now هنا = الحارس العام فقط:
    - النظام مفعّل؟
    - داخل النافذة الزمنية؟
    """
    win = PortalWindow.get()
    msg = (getattr(win, "closed_message", "") or "التقديم مغلق حالياً.").strip()

    if not getattr(win, "is_enabled", True):
        return False, msg, win

    now = timezone.now()
    opens_at = getattr(win, "opens_at", None)
    closes_at = getattr(win, "closes_at", None)

    if opens_at and now < opens_at:
        return False, msg, win

    if closes_at and now > closes_at:
        return False, msg, win

    return True, "", win


def _normalize_portal_phase(value: str) -> str:
    value = (value or "").strip()

    aliases = {
        "closed": "closed",

        "official_only": "official_only",
        "official": "official_only",
        "agents_only": "official_only",

        "new_only": "new_only",
        "new": "new_only",
        "new_applicants_only": "new_only",

        "all": "all",
        "both": "all",
        "all_open": "all",
        "open_all": "all",
    }
    return aliases.get(value, "closed")


def _portal_timer_context(win: PortalWindow) -> dict:
    opens_at = getattr(win, "opens_at", None)
    closes_at = getattr(win, "closes_at", None)
    phase = _normalize_portal_phase(getattr(win, "phase", "closed"))
    is_enabled = getattr(win, "is_enabled", False)

    now = timezone.now()
    now_local = timezone.localtime(now)

    open_by_time = True
    if opens_at and now < opens_at:
        open_by_time = False
    if closes_at and now > closes_at:
        open_by_time = False

    is_portal_open_now = bool(
        is_enabled and phase in {"official_only", "new_only", "all"} and open_by_time
    )

    return {
        "portal_phase": phase,
        "portal_is_enabled": is_enabled,
        "portal_is_open_now": is_portal_open_now,
        "portal_show_countdown": is_portal_open_now and bool(closes_at),
        "portal_opens_at": opens_at,
        "portal_closes_at": closes_at,
        "portal_opens_at_iso": timezone.localtime(opens_at).isoformat() if opens_at else "",
        "portal_closes_at_iso": timezone.localtime(closes_at).isoformat() if closes_at else "",
        "portal_now_iso": now_local.isoformat(),
    }


def _is_official_proxy(applicant: Applicant) -> bool:
    return bool(getattr(applicant, "is_official_agent", False))


def _portal_access_for_applicant(applicant: Applicant, win: PortalWindow):
    special_allowed, special_msg, _special_issue, _special_mode = _special_followup_access_for(applicant)

    open_now, msg, _ = _portal_gate()
    if not open_now:
        if special_allowed:
            return True, special_msg
        return False, special_msg or msg or "التقديم مغلق حالياً."

    phase = _normalize_portal_phase(getattr(win, "phase", "closed"))
    is_official = _is_official_proxy(applicant)

    if phase == "all":
        return True, ""

    if phase == "official_only":
        if not is_official:
            if special_allowed:
                return True, special_msg
            return False, (
                (getattr(win, "official_only_message", "") or "").strip()
                or "التقديم متاح حالياً للوكلاء الرسميين فقط."
            )
        return True, ""

    if phase == "new_only":
        if is_official:
            if special_allowed:
                return True, special_msg
            return False, (
                (getattr(win, "new_only_message", "") or "").strip()
                or "التقديم متاح حالياً للمتقدمين الجدد فقط."
            )
        return True, ""

    if special_allowed:
        return True, special_msg

    return False, (
        special_msg
        or (getattr(win, "closed_message", "") or "").strip()
        or "التقديم مغلق حالياً."
    )

def _eligible_schools_for(applicant: Applicant):
    return (
        SchoolVacancy.objects
        .filter(
            is_open=True,
            sector=applicant.sector,
            gender=applicant.gender,
            reserved_application__isnull=True,
        )
        .exclude(deputy_need=0)
        .order_by("school_name")
    )


def _checked_post(request, name: str) -> bool:
    return (
        (request.POST.get(name) or "").strip().lower()
        in {"1", "true", "on", "yes", "y"}
    )


def _dt_iso(dt) -> str:
    if not dt:
        return ""
    try:
        return timezone.localtime(dt).isoformat()
    except Exception:
        return str(dt)


def _build_submission_snapshot(
    *,
    applicant: Applicant,
    app: Application,
    vacancies: list[SchoolVacancy],
    submitted_at,
    available_count: int,
    preferences_policy_confirmed: bool,
    no_preferences_confirmed: bool,
) -> dict:
    """
    لقطة إثبات محفوظة وقت الإرسال.
    تحفظ نصوص الإقرار وبيانات الرغبات كما كانت وقت الإرسال،
    حتى لو تغيّرت النصوص أو بيانات الشواغر لاحقًا.
    """
    prefs_snapshot = []
    for rank, vacancy in enumerate(vacancies, start=1):
        prefs_snapshot.append({
            "rank": rank,
            "vacancy_id": vacancy.id,
            "school_name": getattr(vacancy, "school_name", "") or "",
            "ministry_no": getattr(vacancy, "ministry_no", "") or "",
            "stage": getattr(vacancy, "stage", "") or "",
            "sector": getattr(vacancy, "sector", "") or "",
            "gender": getattr(vacancy, "gender", "") or "",
            "deputy_need": getattr(vacancy, "deputy_need", 0) or 0,
        })

    return {
        "policy_version": SUBMISSION_POLICY_VERSION,
        "application_id": app.id,
        "submitted_at": _dt_iso(submitted_at),
        "submitted_prefs_count": len(prefs_snapshot),
        "available_count_at_submission": available_count,
        "submitted_without_preferences": len(prefs_snapshot) == 0,
        "no_vacancies_at_submission": available_count == 0,
        "applicant": {
            "id": applicant.id,
            "full_name": getattr(applicant, "full_name", "") or "",
            "national_id": getattr(applicant, "national_id", "") or "",
            "mobile": getattr(applicant, "mobile", "") or "",
            "gender": getattr(applicant, "gender", "") or "",
            "sector": getattr(applicant, "sector", "") or "",
            "rank": getattr(applicant, "rank", "") or "",
            "current_job": getattr(applicant, "current_job", "") or "",
            "current_school": getattr(applicant, "current_school", "") or "",
        },
        "preferences": prefs_snapshot,
        "acknowledgements": {
            "preferences_acknowledged": bool(preferences_policy_confirmed),
            "preferences_ack_text": PREFERENCES_ACK_TEXT if preferences_policy_confirmed else "",
            "preferences_ack_at": _dt_iso(submitted_at) if preferences_policy_confirmed else "",
            "no_preferences_acknowledged": bool(no_preferences_confirmed),
            "no_preferences_ack_text": NO_PREFERENCES_ACK_TEXT if no_preferences_confirmed else "",
            "no_preferences_ack_at": _dt_iso(submitted_at) if no_preferences_confirmed else "",
        },
        "status_after_submission": "submitted",
        "locked_after_submission": True,
        "enters_preference_competition": bool(prefs_snapshot),
        "administrative_processing_eligible": True,
        "competition_meaning": (
            PREFERENCES_COMPETITION_NOTE
            if prefs_snapshot
            else "لا يدخل في مفاضلة الرغبات لعدم تسجيل رغبات."
        ),
        "administrative_meaning": (
            "مستلم للمعالجة ومقفل للتعديل؛ ولا يعني تسجيل الرغبات تحقق التوجيه عليها."
            if prefs_snapshot
            else NO_PREFERENCES_POLICY_MEANING
        ),
    }


def _set_model_field_if_exists(obj, field_name: str, value, update_fields: list[str]):
    """يضبط الحقل فقط إذا كان موجودًا في النموذج؛ لتسهيل الانتقال أثناء التحديث."""
    if any(f.name == field_name for f in obj._meta.fields):
        setattr(obj, field_name, value)
        update_fields.append(field_name)


def _is_final_submission_locked(app: Application | None) -> bool:
    return bool(app and app.locked and app.status == "submitted")


def _is_incomplete_submission_locked(app: Application | None) -> bool:
    """
    إقفال إداري للطلبات غير المكتملة بعد نهاية فترة التقديم.
    لا يغيّر حالة الطلب إلى submitted، ولا ينسب للمتقدم إقرارًا لم يفعله؛
    فقط يمنع استكمال الطلب لاحقًا إذا قررت الإدارة تثبيت الإقفال.
    """
    return bool(app and getattr(app, "locked", False) and getattr(app, "status", "") != "submitted")


def _application_progress_code(app: Application | None) -> str:
    """تصنيف إجرائي لا يعتمد على تغيير قيم status في قاعدة البيانات."""
    if not app:
        return "none"

    if _is_incomplete_submission_locked(app):
        if not getattr(app, "confirmed_at", None):
            return "locked_entered_not_confirmed"
        return "locked_confirmed_not_submitted"

    if not getattr(app, "confirmed_at", None):
        return "entered_not_confirmed"

    if getattr(app, "status", "") != "submitted":
        return "confirmed_not_submitted"

    has_prefs = app.prefs.exists()
    return "submitted_with_prefs" if has_prefs else "submitted_without_prefs"


def _application_progress_label(code: str) -> str:
    labels = {
        "none": "لم يدخل البوابة",
        "entered_not_confirmed": "دخل ولم يؤكد البيانات",
        "confirmed_not_submitted": "أكد ولم يرسل الطلب",
        "submitted_without_prefs": "مرسل بلا رغبات",
        "submitted_with_prefs": "مرسل برغبات",
        "locked_entered_not_confirmed": "مقفل إداريًا: دخل ولم يؤكد",
        "locked_confirmed_not_submitted": "مقفل إداريًا: أكد ولم يرسل",
    }
    return labels.get(code, code or "-")


def _application_progress_note(code: str) -> str:
    notes = {
        "none": "لم يباشر إجراءات التقديم رغم إتاحة البوابة خلال الفترة المحددة.",
        "entered_not_confirmed": "تم إثبات الدخول فقط، ولم يتم تأكيد البيانات.",
        "confirmed_not_submitted": "تم تأكيد البيانات، ولم يتم تنفيذ الإرسال النهائي.",
        "submitted_without_prefs": "تم الإرسال النهائي دون رغبات؛ لا يدخل في مفاضلة الرغبات، مع احتفاظ الإدارة بحق المعالجة وفق المصلحة التعليمية.",
        "submitted_with_prefs": "طلب مكتمل وداخل في مفاضلة الرغبات وفق الضوابط ونتائج المفاضلة والاحتياج.",
        "locked_entered_not_confirmed": "طلب غير مكتمل تم إقفاله إداريًا بعد نهاية فترة التقديم.",
        "locked_confirmed_not_submitted": "طلب غير مكتمل تم إقفاله إداريًا بعد نهاية فترة التقديم.",
    }
    return notes.get(code, "-")



def _application_preferences_count(app: Application, prefs: list[ApplicationPreference] | None = None) -> int:
    """
    يرجع عدد الرغبات بأقل تكلفة ممكنة:
    - يستخدم prefs الجاهزة إن مررت له.
    - يستخدم annotation باسم prefs_count في لوحة الإدارة.
    - وإلا يستعلم من العلاقة.
    """
    if prefs is not None:
        return len(prefs)

    if hasattr(app, "prefs_count"):
        try:
            return int(getattr(app, "prefs_count") or 0)
        except Exception:
            return 0

    try:
        return app.prefs.count()
    except Exception:
        return 0


def _is_submitted_without_preferences(app: Application, prefs: list[ApplicationPreference] | None = None) -> bool:
    """
    مسار إداري مستقل:
    الطلب المرسل دون رغبات لا يدخل مفاضلة الرغبات،
    لكنه يبقى قابلاً للمعالجة الإدارية وفق المصلحة التعليمية.
    """
    if not app or app.status != "submitted":
        return False
    return _application_preferences_count(app, prefs) == 0


def _is_submitted_with_preferences(app: Application, prefs: list[ApplicationPreference] | None = None) -> bool:
    return bool(app and app.status == "submitted" and _application_preferences_count(app, prefs) > 0)


def _admin_decision_display(app: Application, prefs: list[ApplicationPreference] | None = None) -> dict:
    """
    يضبط معنى القرار إداريًا حسب مسار الطلب:
    approved مع رغبات = معتمد، approved بلا رغبات = موثق الاستلام.
    """
    raw = (getattr(app, "admin_decision", "") or "").strip()
    no_prefs = _is_submitted_without_preferences(app, prefs)
    conditional_issue = _application_conditional_data_issue(app)

    if conditional_issue and not raw:
        return {
            "code": "conditional_data_review",
            "label": "معلق على مراجعة البيانات",
            "css": "red",
            "note": "مرسل مشروط؛ لا يعتمد ولا يدخل المفاضلة النهائية حتى تراجع الإدارة طلب تعديل البيانات المؤثر.",
        }

    if raw == "approved" and no_prefs:
        return {
            "code": "documented",
            "label": "موثق الاستلام",
            "css": "green",
            "note": "تم توثيق استلام الطلب دون رغبات، ولا يدخل في مفاضلة الرغبات.",
        }

    if raw == "approved":
        return {
            "code": "approved",
            "label": "معتمد",
            "css": "green",
            "note": "قرار اعتماد ضمن مسار المفاضلة والترشيح.",
        }

    if raw == "rejected":
        return {
            "code": "rejected",
            "label": "مرفوض",
            "css": "red",
            "note": "تم رفض الطلب وفق الملاحظة الإدارية المسجلة.",
        }

    if raw == "returned":
        return {
            "code": "returned",
            "label": "معاد للتعديل",
            "css": "blue",
            "note": "أعيد الطلب للتعديل وفق الملاحظة الإدارية المسجلة.",
        }

    if no_prefs:
        return {
            "code": "pending_documentation",
            "label": "بانتظار توثيق",
            "css": "gold",
            "note": "طلب مرسل دون رغبات وينتظر توثيق الاستلام إداريًا.",
        }

    return {
        "code": "pending",
        "label": "بانتظار قرار",
        "css": "gold",
        "note": "طلب ينتظر اتخاذ القرار الإداري.",
    }


def _application_path_info(app: Application, prefs: list[ApplicationPreference] | None = None) -> dict:
    prefs_count = _application_preferences_count(app, prefs)
    status = (getattr(app, "status", "") or "").strip()
    decision = (getattr(app, "admin_decision", "") or "").strip()
    conditional_issue = _application_conditional_data_issue(app)

    if status == "submitted" and conditional_issue:
        return {
            "code": "conditional_data_review",
            "label": "مرسل مشروط بمراجعة البيانات",
            "status_label": "مرسل مشروط",
            "brief": "تم حفظ الرغبات ووقت الإرسال، والمفاضلة معلقة حتى مراجعة طلب تعديل البيانات.",
            "long_note": "هذا الطلب لا يدخل المفاضلة النهائية ولا يعتمد إداريًا حتى تتم معالجة ملاحظة البيانات المؤثرة.",
            "competition_label": "يدخل المفاضلة النهائية",
            "competition_value": "معلق",
            "admin_processing_label": "مطلوب مراجعة بيانات",
            "admin_processing_value": "نعم",
            "claim_label": "الرغبات محفوظة لحين القرار",
            "claim_value": "محفوظة",
            "primary_action_label": "مراجعة البيانات أولًا",
            "css": "red",
        }

    if status == "submitted" and prefs_count == 0:
        return {
            "code": "no_preferences",
            "label": "مستلم دون رغبات",
            "status_label": "مرسل دون رغبات",
            "brief": "لا يدخل مفاضلة الرغبات — قابل للمعالجة الإدارية",
            "long_note": NO_PREFERENCES_POLICY_MEANING,
            "competition_label": "لا يدخل مفاضلة الرغبات",
            "competition_value": "لا",
            "admin_processing_label": "قابل للمعالجة الإدارية",
            "admin_processing_value": "نعم",
            "claim_label": "لا توجد مطالبة بشاغر محدد",
            "claim_value": "لا",
            "primary_action_label": "توثيق الاستلام",
            "css": "gold",
        }

    if status == "submitted" and prefs_count > 0:
        if decision == "approved":
            label = "معتمد في مسار المفاضلة"
        elif decision == "rejected":
            label = "مرفوض في مسار المفاضلة"
        elif decision == "returned":
            label = "معاد للتعديل"
        else:
            label = "جاهز للمفاضلة"

        return {
            "code": "competition",
            "label": label,
            "status_label": "مرسل برغبات",
            "brief": "يدخل مفاضلة الرغبات وفق الضوابط والدرجة والاحتياج",
            "long_note": PREFERENCES_COMPETITION_NOTE,
            "competition_label": "يدخل مفاضلة الرغبات",
            "competition_value": "نعم",
            "admin_processing_label": "يعالج ضمن مسار المفاضلة",
            "admin_processing_value": "نعم",
            "claim_label": "لا يضمن تحقق الرغبات",
            "claim_value": "لا يضمن",
            "primary_action_label": "اعتماد",
            "css": "green",
        }

    if status == "draft" and getattr(app, "confirmed_at", None):
        return {
            "code": "confirmed_not_submitted",
            "label": "غير مكتمل — أكد ولم يرسل",
            "status_label": "مسودة مؤكدة",
            "brief": "تم تأكيد البيانات ولم يتم الإرسال النهائي.",
            "long_note": "لا يدخل الطلب في المفاضلة قبل الإرسال النهائي.",
            "competition_label": "لا يدخل مفاضلة الرغبات",
            "competition_value": "لا",
            "admin_processing_label": "غير قابل للمعالجة كطلب مرسل",
            "admin_processing_value": "لا",
            "claim_label": "لا توجد مطالبة",
            "claim_value": "لا",
            "primary_action_label": "—",
            "css": "gold",
        }

    if status == "draft":
        return {
            "code": "entered_not_confirmed",
            "label": "غير مكتمل — دخل ولم يؤكد",
            "status_label": "مسودة",
            "brief": "تم إثبات الدخول فقط، ولم يتم تأكيد البيانات أو الإرسال.",
            "long_note": "لا يدخل الطلب في المفاضلة قبل تأكيد البيانات والإرسال النهائي.",
            "competition_label": "لا يدخل مفاضلة الرغبات",
            "competition_value": "لا",
            "admin_processing_label": "غير قابل للمعالجة كطلب مرسل",
            "admin_processing_value": "لا",
            "claim_label": "لا توجد مطالبة",
            "claim_value": "لا",
            "primary_action_label": "—",
            "css": "gold",
        }

    return {
        "code": status or "unknown",
        "label": status or "غير محدد",
        "status_label": status or "غير محدد",
        "brief": "حالة غير مكتملة أو غير محددة.",
        "long_note": "تراجع حالة الطلب وسجل الإجراءات قبل اتخاذ القرار.",
        "competition_label": "غير محدد",
        "competition_value": "—",
        "admin_processing_label": "غير محدد",
        "admin_processing_value": "—",
        "claim_label": "غير محدد",
        "claim_value": "—",
        "primary_action_label": "—",
        "css": "blue",
    }


def _submission_proof_context(app: Application, prefs: list[ApplicationPreference]) -> dict:
    snapshot = getattr(app, "submission_snapshot", None) or {}
    if not isinstance(snapshot, dict):
        snapshot = {}

    prefs_count = _application_preferences_count(app, prefs)
    saved_count = getattr(app, "submitted_prefs_count", None)
    try:
        saved_count = int(saved_count)
    except Exception:
        saved_count = prefs_count

    ack_items = []

    if getattr(app, "preferences_acknowledged", False):
        ack_items.append({
            "title": "إقرار سياسة الرغبات",
            "status": "تم الإقرار",
            "at": _fmt_dt(getattr(app, "preferences_ack_at", None)),
            "text": (getattr(app, "preferences_ack_text", "") or PREFERENCES_ACK_TEXT).strip(),
        })

    if getattr(app, "no_preferences_acknowledged", False):
        ack_items.append({
            "title": "إقرار الإرسال دون رغبات",
            "status": "تم الإقرار",
            "at": _fmt_dt(getattr(app, "no_preferences_ack_at", None)),
            "text": (getattr(app, "no_preferences_ack_text", "") or NO_PREFERENCES_ACK_TEXT).strip(),
        })

    if getattr(app, "status", "") == "submitted" and not ack_items:
        ack_items.append({
            "title": "إثبات الإقرارات",
            "status": "غير محفوظ",
            "at": "—",
            "text": "طلب سابق أو لم يتم العثور على نص إقرار محفوظ ضمن حقول الإثبات الحالية.",
        })

    return {
        "policy_version": getattr(app, "submission_policy_version", "") or SUBMISSION_POLICY_VERSION,
        "submitted_at": _fmt_dt(getattr(app, "submitted_at", None)),
        "submitted_prefs_count": saved_count,
        "snapshot": snapshot,
        "ack_items": ack_items,
        "snapshot_preferences": snapshot.get("preferences", []) if isinstance(snapshot.get("preferences", []), list) else [],
    }


def _enrich_admin_application(app: Application, prefs: list[ApplicationPreference] | None = None) -> Application:
    path_info = _application_path_info(app, prefs)
    decision_info = _admin_decision_display(app, prefs)

    app.path_info = path_info
    app.decision_info = decision_info
    app.path_label = path_info["label"]
    app.path_brief = path_info["brief"]
    app.path_status_label = path_info["status_label"]
    app.path_css = path_info["css"]
    app.is_no_preferences_path = path_info["code"] == "no_preferences"
    app.is_competition_path = path_info["code"] == "competition"
    app.is_conditional_data_review_path = path_info["code"] == "conditional_data_review"
    app.admin_decision_display = decision_info["label"]
    app.admin_decision_css = decision_info["css"]
    app.admin_decision_note_display = decision_info["note"]
    app.primary_action_label = path_info["primary_action_label"]
    return app


def _build_preferences_context(*, applicant: Applicant, app: Application, win: PortalWindow, schools, selected_prefs, selected_ids, error: str = "") -> dict:
    available_count = schools.count()

    # ضابط الاختيار مفتوح:
    # لا يوجد حد أدنى للرغبات، ولا إلزام باختيار جميع الشواغر.
    min_required = 0

    if available_count == 0:
        selection_hint = "لا توجد شواغر متاحة حاليًا في قطاعك، ويمكنك إرسال الطلب دون رغبات."
    else:
        selection_hint = "اختيار الرغبات مفتوح دون حد أدنى إلزامي، ويمكنك اختيار رغبة واحدة أو أكثر حسب رغبتك."

    ctx = {
        "a": applicant,
        "app": app,
        "schools": schools,
        "selected_prefs": selected_prefs,
        "selected_ids": selected_ids,
        "closed_msg": "",
        "available_count": available_count,
        "min_required": min_required,
        "selection_hint": selection_hint,
        "preferences_policy_confirm_text": PREFERENCES_ACK_TEXT,
        "preferences_ack_text": PREFERENCES_ACK_TEXT,
        "no_preferences_ack_text": NO_PREFERENCES_ACK_TEXT,
        "submission_policy_version": SUBMISSION_POLICY_VERSION,
    }
    if error:
        ctx["error"] = error
    ctx.update(_portal_timer_context(win))
    return ctx


def _save_uploaded_file(uploaded_file, prefix: str) -> str:
    os.makedirs(settings.MEDIA_ROOT, exist_ok=True)
    path = os.path.join(settings.MEDIA_ROOT, f"{prefix}__{uploaded_file.name}")
    with open(path, "wb+") as out:
        for chunk in uploaded_file.chunks():
            out.write(chunk)
    return path


def _reset_new_applicants_assignments():
    apps = list(
        Application.objects
        .select_related("applicant", "achieved_pref__vacancy")
        .all()
    )
    apps = [app for app in apps if not _is_official_proxy(app.applicant)]

    released_vacancy_ids: set[int] = set()
    app_ids: list[int] = []

    for app in apps:
        app_ids.append(app.id)
        if (
            getattr(app, "achieved_pref_id", None)
            and getattr(app, "achieved_pref", None)
            and getattr(app.achieved_pref, "vacancy_id", None)
        ):
            released_vacancy_ids.add(app.achieved_pref.vacancy_id)

    if released_vacancy_ids and app_ids:
        SchoolVacancy.objects.filter(
            id__in=released_vacancy_ids,
            reserved_application_id__in=app_ids,
        ).update(
            reserved_application=None,
            reserved_at=None,
        )

    if app_ids:
        Application.objects.filter(id__in=app_ids).update(
            achieved_pref=None,
            achieved_at=None,
            achieved_by=None,
        )


def _run_new_applicants_sorting(*, decided_by):
    base_qs = (
        Application.objects
        .select_related("applicant")
        .prefetch_related("prefs", "prefs__vacancy")
        .filter(
            applicant__is_active=True,
            status="submitted",
        )
        .order_by("submitted_at", "id")
    )

    applications = [
        app for app in base_qs
        if not _is_official_proxy(app.applicant)
        and not _application_is_conditional_data_review(app)
    ]

    if not applications:
        return {
            "applications": 0,
            "assigned": 0,
            "unassigned": 0,
            "available_vacancies": 0,
        }

    _reset_new_applicants_assignments()

    vacancy_map = {
        v.id: v
        for v in SchoolVacancy.objects.filter(
            is_open=True,
            reserved_application__isnull=True,
        ).exclude(deputy_need=0)
    }

    assigned = 0

    for app in applications:
        applicant = app.applicant
        prefs = list(
            app.prefs.select_related("vacancy").order_by("rank", "id")
        )

        matched_pref = None
        for pref in prefs:
            vacancy = getattr(pref, "vacancy", None)
            if not vacancy:
                continue
            if vacancy.id not in vacancy_map:
                continue
            if vacancy.sector != applicant.sector:
                continue
            if vacancy.gender != applicant.gender:
                continue

            matched_pref = pref
            break

        if matched_pref:
            vacancy = matched_pref.vacancy
            vacancy.reserved_application = app
            vacancy.reserved_at = timezone.now()
            # إغلاق تلقائي للشاغر عند تحقق الترشيح؛ ولا يُعاد فتحه إلا يدويًا من إدارة الشواغر.
            vacancy.is_open = False
            vacancy.save(update_fields=["reserved_application", "reserved_at", "is_open"])

            app.achieved_pref = matched_pref
            app.achieved_at = timezone.now()
            app.achieved_by = decided_by
            app.save(update_fields=["achieved_pref", "achieved_at", "achieved_by"])

            vacancy_map.pop(vacancy.id, None)
            assigned += 1

    return {
        "applications": len(applications),
        "assigned": assigned,
        "unassigned": len(applications) - assigned,
        "available_vacancies": len(vacancy_map),
    }


def _is_ajax(request) -> bool:
    return request.headers.get("X-Requested-With") == "XMLHttpRequest"


def _pct(cnt: int, total: int) -> int:
    if not total:
        return 0
    return int(round((cnt * 100) / total))


def _paginate(request, qs, per_page: int = 40):
    paginator = Paginator(qs, per_page)
    page_number = request.GET.get("page") or 1
    return paginator.get_page(page_number)


def _vacancies_filters_querydict(request) -> dict[str, str]:
    params: dict[str, str] = {}
    for key in ("q", "open", "gender", "sector", "page"):
        value = (
            request.POST.get(key)
            or request.GET.get(key)
            or ""
        ).strip()
        if value:
            params[key] = value
    return params


def _redirect_admin_vacancies_list_with_filters(request):
    base_url = redirect("portal:admin_vacancies_list").url
    params = _vacancies_filters_querydict(request)
    if not params:
        return redirect(base_url)
    return redirect(f"{base_url}?{urlencode(params)}")


def _redirect_admin_app_detail_with_back(app_id: int, back_url: str = ""):
    detail_url = redirect("portal:admin_app_detail", app_id=app_id).url
    back_url = (back_url or "").strip()
    if not back_url:
        return redirect(detail_url)
    return redirect(f"{detail_url}?back={quote(back_url, safe='')}")


def _set_admin_decision(app: Application, user, decision: str, note: str):
    decision = (decision or "").strip()

    app.admin_decision = decision
    app.admin_note = (note or "").strip()
    app.admin_decided_by = user
    app.admin_decided_at = timezone.now()

    update_fields = ["admin_decision", "admin_note", "admin_decided_by", "admin_decided_at"]

    if decision in ("rejected", "returned", ""):
        if getattr(app, "achieved_pref_id", None):
            old_pref = app.achieved_pref
            old_vacancy = old_pref.vacancy if old_pref and getattr(old_pref, "vacancy", None) else None

            app.achieved_pref = None
            app.achieved_at = None
            app.achieved_by = None
            update_fields += ["achieved_pref", "achieved_at", "achieved_by"]

            if old_vacancy and old_vacancy.reserved_application_id == app.id:
                old_vacancy.reserved_application = None
                old_vacancy.reserved_at = None
                old_vacancy.save(update_fields=["reserved_application", "reserved_at"])

    app.save(update_fields=update_fields)


def _admin_base_qs():
    return (
        Application.objects
        .select_related("applicant", "achieved_pref__vacancy", "admin_decided_by", "achieved_by")
        .order_by("-submitted_at", "-id")
    )


def _admin_filters_from_request(request):
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()
    sector = (request.GET.get("sector") or "").strip()
    gender = (request.GET.get("gender") or "").strip()
    decision = (request.GET.get("decision") or "").strip()
    return q, status, sector, gender, decision


def _apply_admin_filters(qs, q: str, status: str, sector: str, gender: str, decision: str):
    if status:
        qs = qs.filter(status=status)

    if sector:
        qs = qs.filter(applicant__sector__icontains=sector)

    if gender:
        qs = qs.filter(applicant__gender__icontains=gender)

    if decision:
        if decision == "pending":
            qs = qs.filter(Q(admin_decision__isnull=True) | Q(admin_decision__exact=""))
        else:
            qs = qs.filter(admin_decision=decision)

    if q:
        qs = qs.filter(
            Q(applicant__full_name__icontains=q)
            | Q(applicant__national_id__icontains=q)
            | Q(applicant__sector__icontains=q)
            | Q(applicant__gender__icontains=q)
            | Q(achieved_pref__vacancy__school_name__icontains=q)
            | Q(achieved_pref__vacancy__stage__icontains=q)
        )

    return qs


# =========================================================
# Report Helpers (Nominations)
# =========================================================
def _nominations_filters_from_request(request):
    q = (request.GET.get("q") or "").strip()
    sector = (request.GET.get("sector") or "").strip()
    gender = (request.GET.get("gender") or "").strip()
    school = (request.GET.get("school") or "").strip()
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()
    return q, sector, gender, school, from_date, to_date


def _nominations_qs(request):
    q, sector, gender, school, from_date, to_date = _nominations_filters_from_request(request)

    qs = (
        Application.objects
        .select_related("applicant", "achieved_pref__vacancy", "achieved_by")
        .filter(achieved_pref__isnull=False)
        .order_by("-achieved_at", "-id")
    )

    if q:
        qs = qs.filter(
            Q(applicant__full_name__icontains=q)
            | Q(applicant__national_id__icontains=q)
            | Q(applicant__sector__icontains=q)
            | Q(applicant__gender__icontains=q)
            | Q(achieved_pref__vacancy__school_name__icontains=q)
            | Q(achieved_pref__vacancy__ministry_no__icontains=q)
            | Q(achieved_pref__vacancy__stage__icontains=q)
        )

    if sector:
        qs = qs.filter(applicant__sector__icontains=sector)

    if gender:
        qs = qs.filter(applicant__gender__icontains=gender)

    if school:
        qs = qs.filter(achieved_pref__vacancy__school_name__icontains=school)

    if from_date:
        try:
            dt = datetime.strptime(from_date, "%Y-%m-%d").date()
            qs = qs.filter(achieved_at__date__gte=dt)
        except ValueError:
            pass

    if to_date:
        try:
            dt = datetime.strptime(to_date, "%Y-%m-%d").date()
            qs = qs.filter(achieved_at__date__lte=dt)
        except ValueError:
            pass

    return qs



# =========================================================
# Applicant Data Issues / ملاحظات المتقدمين على البيانات
# =========================================================
DATA_ISSUE_BLOCKING_FIELDS = {
    "gender",
    "rank",
    "sector",
    "current_job",
    "current_school",
    "start_date",
}

DATA_ISSUE_ALLOWED_FIELDS = {
    "full_name",
    "mobile",
    "gender",
    "rank",
    "sector",
    "current_job",
    "current_school",
    "start_date",
    "other",
}

# عند تأخر مراجعة الإدارة لطلب تعديل مؤثر حتى بعد إغلاق البوابة،
# تمنح الإدارة المتقدم مهلة استكمال خاصة بعد صدور القرار.
DATA_ISSUE_FOLLOWUP_WINDOW_HOURS = 24


def _client_ip(request):
    forwarded = request.META.get("HTTP_X_FORWARDED_FOR")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR")


def _applicant_value(applicant, field_name: str) -> str:
    if field_name == "other":
        return ""
    value = getattr(applicant, field_name, "")
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            return str(value)
    return str(value).strip()


def _applicant_snapshot(applicant) -> dict:
    fields = [
        "full_name",
        "national_id",
        "mobile",
        "gender",
        "rank",
        "sector",
        "current_job",
        "current_school",
        "start_date",
    ]
    return {field: _applicant_value(applicant, field) for field in fields}


def _pending_data_issue_for(applicant):
    return (
        ApplicantDataIssue.objects
        .filter(applicant=applicant, status=ApplicantDataIssue.STATUS_PENDING)
        .order_by("-is_blocking", "-created_at", "-id")
        .first()
    )


def _pending_blocking_data_issue_for(applicant):
    return (
        ApplicantDataIssue.objects
        .filter(
            applicant=applicant,
            status=ApplicantDataIssue.STATUS_PENDING,
            is_blocking=True,
        )
        .order_by("-created_at", "-id")
        .first()
    )


def _conditional_data_review_issue_for(applicant):
    """
    طلب تعديل بيانات مؤثر قيد المراجعة.
    لا يمنع المتقدم من تعبئة الرغبات، لكنه يجعل الطلب بعد الإرسال
    مشروطًا بمراجعة البيانات قبل دخوله في المفاضلة النهائية.
    """
    return _pending_blocking_data_issue_for(applicant)


def _application_conditional_data_issue(app: Application | None):
    if not app or not getattr(app, "applicant_id", None):
        return None
    if getattr(app, "status", "") != "submitted":
        return None
    return _conditional_data_review_issue_for(app.applicant)


def _application_is_conditional_data_review(app: Application | None) -> bool:
    return bool(_application_conditional_data_issue(app))


def _conditional_data_issue_snapshot(issue: ApplicantDataIssue | None) -> dict:
    if not issue:
        return {}
    return {
        "issue_id": issue.id,
        "status": issue.status,
        "field_name": issue.field_name,
        "field_label": issue.get_field_name_display(),
        "current_value": issue.current_value or "",
        "proposed_value": issue.proposed_value or "",
        "note": issue.note or "",
        "is_blocking": bool(issue.is_blocking),
        "protects_followup_right": bool(getattr(issue, "protects_followup_right", False)),
        "created_at": _dt_iso(issue.created_at),
        "meaning": (
            "تم السماح للمتقدم باستكمال ترتيب الرغبات حفظًا لحقه، "
            "على أن يبقى الطلب مرسلًا مشروطًا بمراجعة البيانات ولا يدخل المفاضلة النهائية حتى تراجع الإدارة الملاحظة."
        ),
    }


def _pending_protected_data_issue_for(applicant):
    return (
        ApplicantDataIssue.objects
        .filter(
            applicant=applicant,
            status=ApplicantDataIssue.STATUS_PENDING,
            is_blocking=True,
            protects_followup_right=True,
        )
        .order_by("-protected_at", "-created_at", "-id")
        .first()
    )


def _active_followup_data_issue_for(applicant):
    now = timezone.now()
    return (
        ApplicantDataIssue.objects
        .filter(
            applicant=applicant,
            protects_followup_right=True,
            followup_window_expires_at__gte=now,
            status__in=[
                ApplicantDataIssue.STATUS_ALLOWED,
                ApplicantDataIssue.STATUS_CORRECTED,
                ApplicantDataIssue.STATUS_REJECTED,
            ],
        )
        .order_by("-followup_window_expires_at", "-id")
        .first()
    )


def _expired_followup_data_issue_for(applicant):
    now = timezone.now()
    return (
        ApplicantDataIssue.objects
        .filter(
            applicant=applicant,
            protects_followup_right=True,
            followup_window_expires_at__lt=now,
            status__in=[
                ApplicantDataIssue.STATUS_ALLOWED,
                ApplicantDataIssue.STATUS_CORRECTED,
                ApplicantDataIssue.STATUS_REJECTED,
            ],
        )
        .order_by("-followup_window_expires_at", "-id")
        .first()
    )


def _special_followup_access_for(applicant):
    """
    يعيد استثناء الدخول الخاص بطلبات تعديل البيانات المؤثرة:
    - pending_protected: يسمح بالدخول لصفحة التأكيد لمتابعة حالة الملاحظة فقط.
    - active_window: يسمح باستكمال خطوات التقديم خلال مهلة خاصة بعد قرار الإدارة.
    - expired_window: يوضح أن المهلة الخاصة انتهت.
    """
    active_issue = _active_followup_data_issue_for(applicant)
    if active_issue:
        expires = _fmt_dt(active_issue.followup_window_expires_at)
        return True, (
            f"لديك مهلة استكمال خاصة حتى {expires} بسبب طلب تعديل بيانات مؤثر تم رفعه أثناء فترة التقديم."
        ), active_issue, "active_window"

    pending_issue = _pending_protected_data_issue_for(applicant)
    if pending_issue:
        return True, (
            "طلب تعديل بياناتك المؤثر قيد مراجعة الإدارة، وحقك في الاستكمال محفوظ لأنه رُفع أثناء فترة التقديم."
        ), pending_issue, "pending_protected"

    expired_issue = _expired_followup_data_issue_for(applicant)
    if expired_issue:
        return False, (
            "انتهت مهلة الاستكمال الخاصة الممنوحة بعد مراجعة طلب تعديل البيانات."
        ), expired_issue, "expired_window"

    return False, "", None, ""


def _maybe_unlock_for_special_followup(applicant, app: Application | None):
    """يفتح الطلب غير المكتمل إذا كانت للمتقدم مهلة استكمال خاصة نشطة."""
    issue = _active_followup_data_issue_for(applicant)
    if not issue or not app:
        return issue

    if _is_incomplete_submission_locked(app):
        update_fields = []
        if getattr(app, "locked", False):
            app.locked = False
            update_fields.append("locked")
        if getattr(app, "status", "") != "draft":
            app.status = "draft"
            update_fields.append("status")
        if update_fields:
            app.save(update_fields=update_fields)

    return issue


def _grant_followup_window_for_issue(issue: ApplicantDataIssue, user, reason: str = "") -> list[str]:
    """يفتح مهلة استكمال خاصة بعد مراجعة طلب تعديل مؤثر محفوظ الحق."""
    if not getattr(issue, "protects_followup_right", False):
        return []
    if not getattr(issue, "is_blocking", False):
        return []

    now = timezone.now()
    expires_at = now + timedelta(hours=DATA_ISSUE_FOLLOWUP_WINDOW_HOURS)

    issue.followup_window_granted_at = now
    issue.followup_window_expires_at = expires_at
    issue.followup_window_granted_by = user
    issue.followup_window_note = (
        reason
        or f"تم فتح مهلة استكمال خاصة لمدة {DATA_ISSUE_FOLLOWUP_WINDOW_HOURS} ساعة بعد مراجعة طلب تعديل مؤثر."
    )

    return [
        "followup_window_granted_at",
        "followup_window_expires_at",
        "followup_window_granted_by",
        "followup_window_note",
    ]


# =========================================================
# Portal: Closed Page
# =========================================================
@require_GET
def closed_view(request):
    win = PortalWindow.get()
    open_now, msg, _ = _portal_gate()

    a = _get_applicant(request)
    if a:
        allowed, deny_msg = _portal_access_for_applicant(a, win)
        if allowed:
            return redirect("portal:login")
        msg = deny_msg or msg

    if open_now and not a:
        return redirect("portal:login")

    ctx = {"msg": msg}
    ctx.update(_portal_timer_context(win))
    return render(request, "portal/closed.html", ctx)


# =========================================================
# Applicant Portal
# =========================================================
@require_http_methods(["GET", "POST"])
def login_view(request):
    win = PortalWindow.get()

    if request.method == "POST":
        open_now, msg, win = _portal_gate()

        nid = (request.POST.get("national_id") or "").strip().replace(" ", "")

        if not nid:
            ctx = {"error": "فضلاً أدخل السجل المدني"}
            ctx.update(_portal_timer_context(win))
            return render(request, "portal/login.html", ctx)

        if (not nid.isdigit()) or (len(nid) != 10):
            ctx = {"error": "فضلاً أدخل السجل المدني بشكل صحيح"}
            ctx.update(_portal_timer_context(win))
            return render(request, "portal/login.html", ctx)

        applicant = Applicant.objects.filter(national_id=nid, is_active=True).first()
        if not applicant:
            ctx = {"error": "لا يوجد بيانات لهذا السجل المدني."}
            ctx.update(_portal_timer_context(win))
            return render(request, "portal/login.html", ctx)

        allowed, deny_msg = _portal_access_for_applicant(applicant, win)
        if not allowed:
            ctx = {"error": deny_msg}
            ctx.update(_portal_timer_context(win))
            return render(request, "portal/login.html", ctx)

        request.session[SESSION_KEY] = applicant.national_id

        # ضابط إثبات الدخول:
        # بمجرد نجاح دخول المرشح يتم إنشاء طلب Draft له إن لم يكن موجودًا،
        # وبذلك يظهر إداريًا ضمن من دخلوا البوابة حتى لو خرج قبل اختيار الرغبات.
        app, _created = Application.objects.get_or_create(
            applicant=applicant,
            defaults={"status": "draft"},
        )

        if _is_final_submission_locked(app):
            return redirect("portal:done")

        active_followup_issue = _maybe_unlock_for_special_followup(applicant, app)

        if _is_incomplete_submission_locked(app) and not active_followup_issue:
            messages.error(
                request,
                "انتهت فترة التقديم وتم إقفال طلبك غير المكتمل، ولا يمكن استكماله بعد الإغلاق."
            )
            return redirect("portal:closed")

        return redirect("portal:confirm")

    ctx = {}
    ctx.update(_portal_timer_context(win))
    return render(request, "portal/login.html", ctx)


@require_http_methods(["GET", "POST"])
def confirm_view(request):
    a = _get_applicant(request)
    if not a:
        return redirect("portal:login")

    win = PortalWindow.get()
    allowed, deny_msg = _portal_access_for_applicant(a, win)
    if not allowed:
        messages.error(request, deny_msg)
        return redirect("portal:closed")

    app = Application.objects.filter(applicant=a).first()
    active_followup_issue = _maybe_unlock_for_special_followup(a, app)
    if app:
        if _is_final_submission_locked(app):
            messages.info(request, "تم إرسال طلبك مسبقًا ولا يمكن تعديله حالياً.")
            return redirect("portal:done")

        if _is_incomplete_submission_locked(app) and not active_followup_issue:
            messages.error(
                request,
                "انتهت فترة التقديم وتم إقفال طلبك غير المكتمل، ولا يمكن استكماله بعد الإغلاق."
            )
            return redirect("portal:closed")

    def gv(attr: str, dash: str = "-"):
        v = getattr(a, attr, None)
        if v is None:
            return dash
        if isinstance(v, str):
            v = v.strip()
            return v if v else dash
        if hasattr(v, "strftime"):
            try:
                return v.strftime("%Y-%m-%d")
            except Exception:
                return str(v)
        return v

    fields = [
        {"label": "الاسم الرباعي", "value": gv("full_name")},
        {"label": "رقم الهوية", "value": gv("national_id")},
        {"label": "رقم الجوال", "value": gv("mobile")},
        {"label": "الجنس", "value": gv("gender")},
        {"label": "الرتبة", "value": gv("rank")},
        {"label": "القطاع", "value": gv("sector")},
        {"label": "العمل الحالي", "value": gv("current_job")},
        {"label": "المدرسة الحالية", "value": gv("current_school")},
        {"label": "تاريخ المباشرة", "value": gv("start_date")},
    ]

    pending_data_issue = _pending_data_issue_for(a)

    if request.method == "POST":
        app, _ = Application.objects.get_or_create(applicant=a)

        if _is_final_submission_locked(app):
            messages.info(request, "تم إرسال طلبك مسبقًا ولا يمكن تعديله حالياً.")
            return redirect("portal:done")

        form_action = (request.POST.get("form_action") or "confirm_data").strip()

        if form_action == "report_data_issue":
            field_name = (request.POST.get("issue_field") or "").strip()
            proposed_value = (request.POST.get("proposed_value") or "").strip()[:255]
            note = (request.POST.get("issue_note") or "").strip()[:1000]
            confirmed = request.POST.get("issue_confirmed") == "1"

            if field_name not in DATA_ISSUE_ALLOWED_FIELDS:
                messages.error(request, "يلزم اختيار الحقل محل الملاحظة.")
                return redirect("portal:confirm")

            if not note:
                messages.error(request, "يلزم كتابة وصف الملاحظة.")
                return redirect("portal:confirm")

            if not confirmed:
                messages.error(request, "يلزم الإقرار بصحة الملاحظة قبل إرسالها للإدارة.")
                return redirect("portal:confirm")

            is_blocking = field_name in DATA_ISSUE_BLOCKING_FIELDS
            open_now_for_protection, _protection_msg, _protection_win = _portal_gate()
            protects_followup_right = bool(is_blocking and open_now_for_protection)
            now_for_issue = timezone.now()

            issue = ApplicantDataIssue.objects.create(
                applicant=a,
                application=app,
                field_name=field_name,
                current_value=_applicant_value(a, field_name),
                proposed_value=proposed_value,
                note=note,
                is_blocking=is_blocking,
                protects_followup_right=protects_followup_right,
                protected_at=now_for_issue if protects_followup_right else None,
                applicant_snapshot=_applicant_snapshot(a),
                source_ip=_client_ip(request),
                user_agent=(request.META.get("HTTP_USER_AGENT") or "")[:1000],
            )

            if is_blocking:
                messages.warning(
                    request,
                    "تم إرسال ملاحظتك للإدارة. يمكنك المتابعة وترتيب الرغبات حفظًا لحقك، وسيبقى الطلب بعد الإرسال مشروطًا بمراجعة البيانات قبل دخوله في المفاضلة النهائية."
                )
            else:
                messages.success(
                    request,
                    "تم إرسال ملاحظتك للإدارة، ويمكنك المتابعة مع بقاء الملاحظة موثقة وقيد المراجعة."
                )

            ctx = {
                "a": a,
                "fields": fields,
                "pending_data_issue": issue,
                "data_issue_created": True,
                "issue_blocks_followup": is_blocking,
                "active_followup_issue": active_followup_issue,
                "pending_protected_issue": issue if protects_followup_right else None,
            }
            ctx.update(_portal_timer_context(win))
            return render(request, "portal/confirm.html", ctx)

        # تأكيد البيانات والمتابعة:
        # إذا توجد ملاحظة مؤثرة قيد المراجعة، لا نوقف تعبئة الرغبات.
        # يسجَّل الطلب لاحقًا كطلب مشروط بمراجعة البيانات قبل المفاضلة النهائية.
        pending_data_issue = _pending_data_issue_for(a)
        if pending_data_issue and pending_data_issue.is_blocking:
            messages.warning(
                request,
                "سيتم السماح لك بترتيب الرغبات حفظًا لحقك، وسيبقى الإرسال مشروطًا بمراجعة الإدارة لطلب تعديل البيانات."
            )

        update_fields: list[str] = []
        if not app.confirmed_at:
            app.confirmed_at = timezone.now()
            update_fields.append("confirmed_at")

        if app.status != "draft":
            app.status = "draft"
            update_fields.append("status")

        if getattr(app, "locked", False):
            app.locked = False
            update_fields.append("locked")

        if update_fields:
            app.save(update_fields=update_fields)

        return redirect("portal:preferences")

    ctx = {
        "a": a,
        "fields": fields,
        "pending_data_issue": pending_data_issue,
        "active_followup_issue": active_followup_issue,
        "pending_protected_issue": _pending_protected_data_issue_for(a),
        "expired_followup_issue": _expired_followup_data_issue_for(a),
    }
    ctx.update(_portal_timer_context(win))
    return render(request, "portal/confirm.html", ctx)


def preferences_view(request):
    a = _get_applicant(request)
    if not a:
        return redirect("portal:login")

    win = PortalWindow.get()
    allowed, deny_msg = _portal_access_for_applicant(a, win)
    if not allowed:
        messages.error(request, deny_msg)
        return redirect("portal:closed")

    app = Application.objects.filter(applicant=a).first()
    if not app or not app.confirmed_at:
        messages.info(request, "يلزم تأكيد البيانات أولاً قبل إدخال الرغبات.")
        return redirect("portal:confirm")

    active_followup_issue = _maybe_unlock_for_special_followup(a, app)
    pending_data_issue = _pending_data_issue_for(a)
    conditional_data_issue = _conditional_data_review_issue_for(a)

    if _is_final_submission_locked(app):
        return redirect("portal:done")

    if _is_incomplete_submission_locked(app) and not active_followup_issue:
        messages.error(
            request,
            "انتهت فترة التقديم وتم إقفال طلبك غير المكتمل، ولا يمكن استكماله بعد الإغلاق."
        )
        return redirect("portal:closed")

    selected_prefs = list(
        ApplicationPreference.objects
        .filter(application=app)
        .select_related("vacancy")
        .order_by("rank", "id")
    )
    selected_ids = [p.vacancy_id for p in selected_prefs]
    schools = _eligible_schools_for(a)

    ctx = _build_preferences_context(
        applicant=a,
        app=app,
        win=win,
        schools=schools,
        selected_prefs=selected_prefs,
        selected_ids=selected_ids,
    )
    ctx["pending_data_issue"] = pending_data_issue
    ctx["conditional_data_issue"] = conditional_data_issue
    ctx["is_conditional_data_review"] = bool(conditional_data_issue)
    return render(request, "portal/preferences.html", ctx)


@transaction.atomic
@require_POST
def submit_view(request):
    a = _get_applicant(request)
    if not a:
        return redirect("portal:login")

    app = get_object_or_404(Application, applicant=a)

    win = PortalWindow.get()
    allowed, deny_msg = _portal_access_for_applicant(a, win)
    if not allowed:
        messages.error(request, deny_msg)
        return redirect("portal:closed")

    if not app.confirmed_at:
        messages.error(request, "يلزم تأكيد البيانات أولاً قبل إرسال الرغبات.")
        return redirect("portal:confirm")

    active_followup_issue = _maybe_unlock_for_special_followup(a, app)
    pending_data_issue = _pending_data_issue_for(a)
    conditional_data_issue = _conditional_data_review_issue_for(a)

    if _is_final_submission_locked(app):
        messages.info(request, "تم إرسال طلبك مسبقًا ولا يمكن تعديله حالياً.")
        return redirect("portal:done")

    if _is_incomplete_submission_locked(app) and not active_followup_issue:
        messages.error(
            request,
            "انتهت فترة التقديم وتم إقفال طلبك غير المكتمل، ولا يمكن استكماله بعد الإغلاق."
        )
        return redirect("portal:closed")

    ids = request.POST.getlist("vacancy_ids")

    schools = _eligible_schools_for(a)
    allowed_ids = set(schools.values_list("id", flat=True))
    available_count = len(allowed_ids)

    clean_ids: list[int] = []
    for x in ids:
        try:
            vid = int(x)
        except Exception:
            continue

        if vid in allowed_ids and vid not in clean_ids:
            clean_ids.append(vid)

    # ضابط الإرسال بدون رغبات:
    # الرغبات غير إلزامية، لكن إذا كانت هناك شواغر متاحة ولم يختر المرشح أي رغبة،
    # فلا يتم الإرسال النهائي إلا بعد إقرار صريح منه بأنه اطلع ويرغب بالإرسال دون رغبات.
    no_preferences_confirmed = _checked_post(request, "confirm_no_preferences")

    # ضابط إقرار سياسة التوجيه عند اختيار رغبات:
    # لا يُقبل إرسال الطلب برغبات مختارة إلا بعد إقرار صريح بأن التوجيه
    # لا يعني الاستحقاق المباشر للرغبات، وإنما يكون وفق المصلحة التعليمية
    # واحتياج الإدارة والضوابط المعتمدة، وفي حدود الرغبات المحددة.
    preferences_policy_confirmed = _checked_post(request, "confirm_preferences_policy")

    if clean_ids and not preferences_policy_confirmed:
        selected_prefs = list(
            ApplicationPreference.objects
            .filter(application=app)
            .select_related("vacancy")
            .order_by("rank", "id")
        )
        selected_ids = [p.vacancy_id for p in selected_prefs]

        ctx = _build_preferences_context(
            applicant=a,
            app=app,
            win=win,
            schools=schools,
            selected_prefs=selected_prefs,
            selected_ids=selected_ids,
            error=(
                "يلزم تفعيل إقرار سياسة التوجيه قبل إرسال الرغبات؛ "
                "فاختيار الرغبات لا يعني تحقق التوجيه عليها، "
                "ولا يمنح أولوية على من هو أعلى درجة أو أحق في المفاضلة، "
                "ويكون التوجيه وفق المصلحة التعليمية واحتياج الإدارة "
                "والضوابط المعتمدة ونتائج المفاضلة."
            ),
        )
        return render(request, "portal/preferences.html", ctx)

    if available_count > 0 and not clean_ids and not no_preferences_confirmed:
        selected_prefs = list(
            ApplicationPreference.objects
            .filter(application=app)
            .select_related("vacancy")
            .order_by("rank", "id")
        )
        selected_ids = [p.vacancy_id for p in selected_prefs]

        ctx = _build_preferences_context(
            applicant=a,
            app=app,
            win=win,
            schools=schools,
            selected_prefs=selected_prefs,
            selected_ids=selected_ids,
            error=(
                "لم تقم باختيار أي رغبة. إذا كنت ترغب في إرسال الطلب دون رغبات، "
                "فضلاً فعّل إقرار الإرسال دون رغبات؛ وسيُعد الطلب مستلمًا دون رغبات، "
                "ولا يدخل في مفاضلة الرغبات مع احتفاظ الإدارة بحق المعالجة وفق المصلحة التعليمية."
            ),
        )
        ctx["require_no_preferences_confirm"] = True
        ctx["no_preferences_confirm_text"] = NO_PREFERENCES_ACK_TEXT
        return render(request, "portal/preferences.html", ctx)

    # نحفظ الشواغر المختارة بترتيب clean_ids قبل إنشاء الرغبات؛
    # حتى تدخل نفس البيانات في لقطة الإرسال المحفوظة.
    selected_vacancy_map = {
        v.id: v
        for v in SchoolVacancy.objects.filter(id__in=clean_ids)
    }
    selected_vacancies = [
        selected_vacancy_map[vid]
        for vid in clean_ids
        if vid in selected_vacancy_map
    ]

    submitted_at = timezone.now()

    ApplicationPreference.objects.filter(application=app).delete()

    for idx, vacancy in enumerate(selected_vacancies, start=1):
        ApplicationPreference.objects.create(
            application=app,
            vacancy=vacancy,
            rank=idx,
        )

    # لا نسجل إقرار عدم اختيار الرغبات إلا إذا فعّله المتقدم صراحة.
    # أما إذا لم تكن هناك شواغر متاحة أصلًا، فتسجل لقطة الإرسال ذلك دون نسبة إقرار غير موجود.
    no_preferences_acknowledged = bool(not selected_vacancies and no_preferences_confirmed)
    preferences_acknowledged = bool(selected_vacancies and preferences_policy_confirmed)

    snapshot = _build_submission_snapshot(
        applicant=a,
        app=app,
        vacancies=selected_vacancies,
        submitted_at=submitted_at,
        available_count=available_count,
        preferences_policy_confirmed=preferences_acknowledged,
        no_preferences_confirmed=no_preferences_acknowledged,
    )

    if conditional_data_issue:
        snapshot["conditional_data_review"] = _conditional_data_issue_snapshot(conditional_data_issue)
        snapshot["enters_preference_competition"] = False
        snapshot["competition_meaning"] = (
            "مرسل مشروط بمراجعة البيانات؛ لا يدخل المفاضلة النهائية حتى تتم معالجة طلب تعديل البيانات المؤثر."
        )
        snapshot["administrative_meaning"] = (
            "تم حفظ وقت الإرسال والرغبات، مع تعليق المفاضلة النهائية إلى حين مراجعة الإدارة لطلب تعديل البيانات."
        )

    app.status = "submitted"
    app.locked = True
    app.submitted_at = submitted_at

    update_fields = ["status", "locked", "submitted_at"]

    # عند إعادة إرسال الطلب بعد إرجاعه للتعديل:
    # نمسح قرار الإرجاع السابق حتى يعود الطلب إلى انتظار القرار،
    # فيظهر إداريًا بعبارة (جاهز للمفاضلة) ومع زر (اعتماد).
    if (getattr(app, "admin_decision", "") or "").strip() == "returned":
        app.admin_decision = ""
        app.admin_note = ""
        app.admin_decided_by = None
        app.admin_decided_at = None
        update_fields += [
            "admin_decision",
            "admin_note",
            "admin_decided_by",
            "admin_decided_at",
        ]

    _set_model_field_if_exists(app, "preferences_acknowledged", preferences_acknowledged, update_fields)
    _set_model_field_if_exists(
        app,
        "preferences_ack_text",
        PREFERENCES_ACK_TEXT if preferences_acknowledged else "",
        update_fields,
    )
    _set_model_field_if_exists(
        app,
        "preferences_ack_at",
        submitted_at if preferences_acknowledged else None,
        update_fields,
    )

    _set_model_field_if_exists(app, "no_preferences_acknowledged", no_preferences_acknowledged, update_fields)
    _set_model_field_if_exists(
        app,
        "no_preferences_ack_text",
        NO_PREFERENCES_ACK_TEXT if no_preferences_acknowledged else "",
        update_fields,
    )
    _set_model_field_if_exists(
        app,
        "no_preferences_ack_at",
        submitted_at if no_preferences_acknowledged else None,
        update_fields,
    )

    _set_model_field_if_exists(app, "submitted_prefs_count", len(selected_vacancies), update_fields)
    _set_model_field_if_exists(app, "submission_policy_version", SUBMISSION_POLICY_VERSION, update_fields)
    _set_model_field_if_exists(app, "submission_snapshot", snapshot, update_fields)

    app.save(update_fields=update_fields)

    return redirect("portal:done")


def done_view(request):
    a = _get_applicant(request)
    if not a:
        return redirect("portal:login")

    win = PortalWindow.get()

    app = (
        Application.objects
        .select_related("applicant", "achieved_pref__vacancy")
        .prefetch_related("prefs", "prefs__vacancy")
        .filter(applicant=a)
        .first()
    )
    prefs = list(app.prefs.select_related("vacancy").all()) if app else []
    no_vacancies = bool(app and app.status == "submitted" and not prefs)

    ctx = {
        "a": a,
        "app": app,
        "prefs": prefs,
        "no_vacancies": no_vacancies,
    }
    ctx.update(_portal_timer_context(win))
    return render(request, "portal/done.html", ctx)



# =========================================================
# Admin: Applicant Data Issues
# =========================================================
@staff_member_required
def admin_data_issues_view(request):
    q = (request.GET.get("q") or "").strip()
    mode = (request.GET.get("mode") or "pending").strip()
    status = (request.GET.get("status") or "").strip()
    blocking = (request.GET.get("blocking") or "").strip()

    now = timezone.now()

    qs = (
        ApplicantDataIssue.objects
        .select_related("applicant", "application", "reviewed_by")
        .annotate(
            review_priority=Case(
                When(status=ApplicantDataIssue.STATUS_PENDING, is_blocking=True, then=Value(1)),
                When(
                    protects_followup_right=True,
                    followup_window_expires_at__gte=now,
                    status__in=[
                        ApplicantDataIssue.STATUS_ALLOWED,
                        ApplicantDataIssue.STATUS_CORRECTED,
                        ApplicantDataIssue.STATUS_REJECTED,
                    ],
                    then=Value(2),
                ),
                When(status=ApplicantDataIssue.STATUS_PENDING, is_blocking=False, then=Value(3)),
                default=Value(4),
                output_field=IntegerField(),
            )
        )
        .order_by("review_priority", "-created_at", "-id")
    )

    # فلاتر عملية مبنية على السؤال: ما المطلوب من الإدارة الآن؟
    if mode == "blocking":
        qs = qs.filter(status=ApplicantDataIssue.STATUS_PENDING, is_blocking=True)
        status = status or ApplicantDataIssue.STATUS_PENDING
        blocking = blocking or "yes"
    elif mode == "active_window":
        qs = qs.filter(
            protects_followup_right=True,
            followup_window_expires_at__gte=now,
            status__in=[
                ApplicantDataIssue.STATUS_ALLOWED,
                ApplicantDataIssue.STATUS_CORRECTED,
                ApplicantDataIssue.STATUS_REJECTED,
            ],
        )
    elif mode == "processed":
        qs = qs.exclude(status=ApplicantDataIssue.STATUS_PENDING)
    elif mode == "all":
        pass
    else:
        mode = "pending"
        qs = qs.filter(status=ApplicantDataIssue.STATUS_PENDING)
        status = status or ApplicantDataIssue.STATUS_PENDING

    # توافق مع الفلاتر التفصيلية اليدوية.
    if status and status != "all":
        qs = qs.filter(status=status)

    if blocking == "yes":
        qs = qs.filter(is_blocking=True)
    elif blocking == "no":
        qs = qs.filter(is_blocking=False)

    if q:
        qs = qs.filter(
            Q(applicant__full_name__icontains=q)
            | Q(applicant__national_id__icontains=q)
            | Q(applicant__mobile__icontains=q)
            | Q(note__icontains=q)
            | Q(proposed_value__icontains=q)
            | Q(current_value__icontains=q)
        )

    page_obj = _paginate(request, qs, per_page=40)

    active_statuses = [
        ApplicantDataIssue.STATUS_ALLOWED,
        ApplicantDataIssue.STATUS_CORRECTED,
        ApplicantDataIssue.STATUS_REJECTED,
    ]

    ctx = {
        "rows": page_obj,
        "q": q,
        "mode": mode,
        "status": status,
        "blocking": blocking,
        "pending_count": ApplicantDataIssue.objects.filter(status=ApplicantDataIssue.STATUS_PENDING).count(),
        "blocking_count": ApplicantDataIssue.objects.filter(
            status=ApplicantDataIssue.STATUS_PENDING,
            is_blocking=True,
        ).count(),
        "nonblocking_pending_count": ApplicantDataIssue.objects.filter(
            status=ApplicantDataIssue.STATUS_PENDING,
            is_blocking=False,
        ).count(),
        "active_followup_count": ApplicantDataIssue.objects.filter(
            protects_followup_right=True,
            followup_window_expires_at__gte=now,
            status__in=active_statuses,
        ).count(),
        "processed_count": ApplicantDataIssue.objects.exclude(status=ApplicantDataIssue.STATUS_PENDING).count(),
        "status_choices": ApplicantDataIssue.STATUS_CHOICES,
        "followup_window_hours": DATA_ISSUE_FOLLOWUP_WINDOW_HOURS,
    }
    return render(request, "portal/admin_data_issues.html", ctx)


@staff_member_required
@require_POST
def admin_data_issue_review_view(request, pk: int):
    issue = get_object_or_404(
        ApplicantDataIssue.objects.select_related("applicant"),
        pk=pk,
    )
    action = (request.POST.get("action") or "").strip()
    admin_note = (request.POST.get("admin_note") or "").strip()

    if action == "allow":
        # السماح بالمتابعة دون تعديل البيانات.
        issue.status = ApplicantDataIssue.STATUS_ALLOWED
        issue.admin_note = admin_note or (
            "تمت مراجعة الملاحظة والسماح للمتقدم بالمتابعة دون تعديل بياناته؛ "
            "ولا يعد ذلك قبولًا للتصحيح المقترح."
        )

    elif action == "correct":
        # اعتماد التصحيح = تعديل بيانات المتقدم فعليًا بالقيمة المعتمدة من الإدارة.
        if issue.field_name == "other":
            messages.error(
                request,
                "لا يمكن اعتماد التصحيح التلقائي لحقل (أخرى). استخدم السماح بالمتابعة دون تعديل أو الرفض مع ملاحظة إدارية."
            )
            return redirect("portal:admin_data_issues")

        approved_value = (request.POST.get("approved_value") or "").strip()
        if not approved_value:
            approved_value = (issue.proposed_value or "").strip()

        if not approved_value:
            messages.error(
                request,
                "يلزم إدخال القيمة المعتمدة من الإدارة قبل تحديث بيانات المتقدم."
            )
            return redirect("portal:admin_data_issues")

        applicant_field_names = {f.name for f in issue.applicant._meta.fields}
        if issue.field_name not in applicant_field_names:
            messages.error(request, "الحقل المطلوب تصحيحه غير موجود في بيانات المتقدم.")
            return redirect("portal:admin_data_issues")

        old_value = getattr(issue.applicant, issue.field_name, "") or ""
        setattr(issue.applicant, issue.field_name, approved_value)
        update_fields = [issue.field_name]
        if "updated_at" in applicant_field_names:
            update_fields.append("updated_at")
        issue.applicant.save(update_fields=update_fields)

        issue.status = ApplicantDataIssue.STATUS_CORRECTED

        audit_note = (
            f"تم اعتماد التصحيح وتحديث بيانات المتقدم في حقل: {issue.get_field_name_display()}.\n"
            f"القيمة السابقة: {old_value or '—'}\n"
            f"التصحيح المقترح من المتقدم: {(issue.proposed_value or '—')}\n"
            f"القيمة المعتمدة من الإدارة: {approved_value}"
        )
        issue.admin_note = f"{admin_note}\n\n{audit_note}".strip() if admin_note else audit_note

    elif action == "reject":
        # رفض الملاحظة = لا تعديل على بيانات المتقدم، مع توثيق سبب الإدارة.
        issue.status = ApplicantDataIssue.STATUS_REJECTED
        issue.admin_note = admin_note or "تم رفض الملاحظة بعد المراجعة، وتبقى البيانات الحالية معتمدة."

    elif action == "accept":
        # توافق خلفي فقط للزر القديم إن كان موجودًا في متصفح أو قالب غير محدث.
        issue.status = ApplicantDataIssue.STATUS_ALLOWED
        issue.admin_note = admin_note or (
            "تمت مراجعة الملاحظة والسماح بالمتابعة دون تعديل بيانات المتقدم."
        )

    else:
        messages.error(request, "إجراء غير معروف.")
        return redirect("portal:admin_data_issues")

    followup_update_fields = _grant_followup_window_for_issue(
        issue,
        request.user,
        reason=(
            f"تم فتح مهلة استكمال خاصة لمدة {DATA_ISSUE_FOLLOWUP_WINDOW_HOURS} ساعة بعد قرار الإدارة: {issue.get_status_display()}."
        ),
    )
    if followup_update_fields:
        window_note = (
            f"تم حفظ حق المتقدم في الاستكمال، وفتحت له مهلة خاصة حتى "
            f"{_fmt_dt(issue.followup_window_expires_at)}."
        )
        issue.admin_note = f"{issue.admin_note}\n\n{window_note}".strip()

    issue.reviewed_at = timezone.now()
    issue.reviewed_by = request.user
    update_fields = ["status", "admin_note", "reviewed_at", "reviewed_by"] + followup_update_fields
    issue.save(update_fields=update_fields)
    if followup_update_fields:
        messages.success(request, "تم تحديث القرار وفتح مهلة استكمال خاصة للمتقدم.")
    else:
        messages.success(request, "تم تحديث قرار مراجعة ملاحظة البيانات.")
    return redirect("portal:admin_data_issues")


# =========================================================
# Admin: Portal Window (Open/Close)
# =========================================================
@staff_member_required
@require_http_methods(["GET", "POST"])
def admin_portal_window_view(request):
    win = PortalWindow.get()

    if request.method == "POST":
        win.is_enabled = (request.POST.get("is_enabled") == "1")
        win.phase = _normalize_portal_phase(request.POST.get("phase") or "closed")

        opens_at = (request.POST.get("opens_at") or "").strip()
        closes_at = (request.POST.get("closes_at") or "").strip()

        win.closed_message = (
            (request.POST.get("closed_message") or "").strip()
            or "التقديم مغلق حالياً."
        )

        if hasattr(win, "official_only_message"):
            win.official_only_message = (
                (request.POST.get("official_only_message") or "").strip()
                or "التقديم متاح حالياً للوكلاء الرسميين فقط."
            )

        if hasattr(win, "new_only_message"):
            win.new_only_message = (
                (request.POST.get("new_only_message") or "").strip()
                or "التقديم متاح حالياً للمتقدمين الجدد فقط."
            )

        if hasattr(win, "all_message"):
            win.all_message = (
                (request.POST.get("all_message") or "").strip()
                or "التقديم متاح حالياً للجميع."
            )

        def parse_dt(v: str):
            if not v:
                return None
            try:
                naive = datetime.strptime(v, "%Y-%m-%dT%H:%M")
            except Exception:
                return None
            tz = timezone.get_current_timezone()
            return timezone.make_aware(naive, tz)

        win.opens_at = parse_dt(opens_at)
        win.closes_at = parse_dt(closes_at)
        win.save()

        messages.success(request, "تم حفظ إعدادات فترة التقديم.")
        return redirect("portal:admin_portal_window")

    ctx = {"win": win}
    ctx.update(_portal_timer_context(win))
    return render(request, "portal/admin_portal_window.html", ctx)


# =========================================================
# Admin: Manage Applicants
# =========================================================
@staff_member_required
def admin_applicants_list(request):
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()
    kind = (request.GET.get("kind") or "all").strip()

    qs = Applicant.objects.all().order_by("-id")

    if status == "active":
        qs = qs.filter(is_active=True)
    elif status == "inactive":
        qs = qs.filter(is_active=False)

    if q:
        qs = qs.filter(
            Q(full_name__icontains=q)
            | Q(national_id__icontains=q)
            | Q(sector__icontains=q)
            | Q(mobile__icontains=q)
            | Q(current_school__icontains=q)
            | Q(current_job__icontains=q)
        )

    all_rows = list(qs)
    official_rows = [a for a in all_rows if a.is_official_agent]
    new_rows = [a for a in all_rows if a.is_new_applicant]

    if kind == "official":
        filtered_rows = official_rows
    elif kind == "new":
        filtered_rows = new_rows
    else:
        filtered_rows = all_rows

    page_obj = _paginate(request, filtered_rows, per_page=40)

    return render(
        request,
        "portal/admin_applicants_list.html",
        {
            "rows": page_obj,
            "q": q,
            "status": status,
            "kind": kind,
            "total": len(filtered_rows),
            "all_count": len(all_rows),
            "official_count": len(official_rows),
            "new_count": len(new_rows),
        },
    )


@staff_member_required
def admin_applicants_create(request):
    form = ApplicantAdminForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تم إضافة المتقدم.")
        return redirect("portal:admin_applicants_list")
    return render(request, "portal/admin_applicants_form.html", {"form": form, "mode": "create"})


@staff_member_required
def admin_applicants_edit(request, pk: int):
    obj = get_object_or_404(Applicant, pk=pk)
    form = ApplicantAdminForm(request.POST or None, instance=obj)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تم حفظ التعديل.")
        return redirect("portal:admin_applicants_list")

    apps_count = Application.objects.filter(applicant=obj).count()
    return render(
        request,
        "portal/admin_applicants_form.html",
        {"form": form, "mode": "edit", "obj": obj, "apps_count": apps_count},
    )


@staff_member_required
@require_POST
def admin_applicants_toggle(request, pk: int):
    obj = get_object_or_404(Applicant, pk=pk)
    obj.is_active = not obj.is_active
    obj.save(update_fields=["is_active"])
    messages.success(request, "تم تحديث حالة المتقدم.")
    return redirect("portal:admin_applicants_list")


@staff_member_required
@require_POST
def admin_applicants_disable_all_view(request):
    updated = Applicant.objects.filter(is_active=True).update(is_active=False)
    messages.success(request, f"تم تعطيل جميع المتقدمين بنجاح. العدد المتأثر: {updated}")
    return redirect("portal:admin_applicants_list")


@staff_member_required
@require_POST
def admin_applicants_enable_all_view(request):
    updated = Applicant.objects.filter(is_active=False).update(is_active=True)
    messages.success(request, f"تم تفعيل جميع المتقدمين بنجاح. العدد المتأثر: {updated}")
    return redirect("portal:admin_applicants_list")



@staff_member_required
@require_POST
def admin_applicants_bulk_action_view(request):
    """
    إجراء جماعي على المتقدمين المحددين من صفحة إدارة المتقدمين:
    - enable  : تفعيل المحدد
    - disable : تعطيل المحدد
    """
    selected_ids = request.POST.getlist("selected_applicants")
    bulk_action = (request.POST.get("bulk_action") or "").strip()

    if not selected_ids:
        messages.warning(request, "لم يتم تحديد أي متقدم.")
        return redirect("portal:admin_applicants_list")

    qs = Applicant.objects.filter(id__in=selected_ids)

    if bulk_action == "enable":
        updated = qs.update(is_active=True)
        messages.success(request, f"تم تفعيل ({updated}) من المتقدمين المحددين.")
    elif bulk_action == "disable":
        updated = qs.update(is_active=False)
        messages.success(request, f"تم تعطيل ({updated}) من المتقدمين المحددين.")
    else:
        messages.warning(request, "إجراء غير معروف.")

    return redirect("portal:admin_applicants_list")


@staff_member_required
@require_POST
def admin_applicants_delete(request, pk: int):
    if not request.user.is_superuser:
        messages.error(request, "غير مصرح بالحذف النهائي. استخدم التعطيل.")
        return redirect("portal:admin_applicants_list")

    obj = get_object_or_404(Applicant, pk=pk)

    if Application.objects.filter(applicant=obj).exists():
        messages.error(request, "لا يمكن الحذف النهائي: المتقدم لديه طلبات مرتبطة. استخدم التعطيل بدلًا من ذلك.")
        return redirect("portal:admin_applicants_list")

    obj.delete()
    messages.success(request, "تم حذف المتقدم نهائيًا.")
    return redirect("portal:admin_applicants_list")


# =========================================================
# Admin: Manage Vacancies + Counts
# =========================================================
@staff_member_required
def admin_vacancies_list(request):
    q = (request.GET.get("q") or "").strip()
    open_state = (request.GET.get("open") or "").strip()
    gender = (request.GET.get("gender") or "").strip()
    sector = (request.GET.get("sector") or "").strip()

    achieved_sq = (
        Application.objects
        .filter(achieved_pref__vacancy_id=OuterRef("pk"))
        .values("achieved_pref__vacancy_id")
        .annotate(c=Count("id"))
        .values("c")[:1]
    )

    # مؤشر التفعيل الجزئي:
    # إذا كانت المدرسة مكررة كسجلات شواغر، وكان بعض سجلاتها مفتوحًا وبعضها مغلقًا،
    # يظهر مربع/مؤشر ذهبي بجوار اسم المدرسة في القالب.
    same_school_total_sq = (
        SchoolVacancy.objects
        .filter(
            school_name=OuterRef("school_name"),
            sector=OuterRef("sector"),
            gender=OuterRef("gender"),
        )
        .order_by()
        .values("school_name", "sector", "gender")
        .annotate(c=Count("id"))
        .values("c")[:1]
    )

    same_school_open_sq = (
        SchoolVacancy.objects
        .filter(
            school_name=OuterRef("school_name"),
            sector=OuterRef("sector"),
            gender=OuterRef("gender"),
            is_open=True,
        )
        .order_by()
        .values("school_name", "sector", "gender")
        .annotate(c=Count("id"))
        .values("c")[:1]
    )

    qs = (
        SchoolVacancy.objects
        .all()
        .annotate(
            interested_total=Count("applicationpreference", distinct=True),
            interested_rank1=Count(
                "applicationpreference",
                filter=Q(applicationpreference__rank=1),
                distinct=True,
            ),
            achieved_total=Coalesce(Subquery(achieved_sq, output_field=IntegerField()), Value(0)),
            school_vacancy_count=Coalesce(Subquery(same_school_total_sq, output_field=IntegerField()), Value(1)),
            school_open_count=Coalesce(Subquery(same_school_open_sq, output_field=IntegerField()), Value(0)),
        )
        .order_by("-id")
    )

    if open_state == "open":
        qs = qs.filter(is_open=True)
    elif open_state == "closed":
        qs = qs.filter(is_open=False)

    if gender:
        qs = qs.filter(gender__icontains=gender)

    if sector:
        qs = qs.filter(sector__icontains=sector)

    if q:
        qs = qs.filter(
            Q(school_name__icontains=q)
            | Q(ministry_no__icontains=q)
            | Q(sector__icontains=q)
            | Q(manager_name__icontains=q)
            | Q(manager_national_id__icontains=q)
        )

    page_obj = _paginate(request, qs, per_page=40)

    # نضيف الخاصية على كائنات الصفحة حتى يستخدمها القالب مباشرة.
    for v in page_obj.object_list:
        total_for_school = int(getattr(v, "school_vacancy_count", 0) or 0)
        open_for_school = int(getattr(v, "school_open_count", 0) or 0)
        v.is_partial_activation = bool(
            total_for_school > 1
            and open_for_school > 0
            and open_for_school < total_for_school
        )

    return render(
        request,
        "portal/admin_vacancies_list.html",
        {
            "rows": page_obj,
            "q": q,
            "open": open_state,
            "gender": gender,
            "sector": sector,
            "total": qs.count(),
        },
    )


@staff_member_required
def admin_vacancies_create(request):
    form = VacancyAdminForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تم إضافة الشاغر/المدرسة.")
        return redirect("portal:admin_vacancies_list")
    return render(request, "portal/admin_vacancies_form.html", {"form": form, "mode": "create"})


@staff_member_required
def admin_vacancies_edit(request, pk: int):
    obj = get_object_or_404(SchoolVacancy, pk=pk)
    form = VacancyAdminForm(request.POST or None, instance=obj)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تم حفظ التعديل.")
        return redirect("portal:admin_vacancies_list")

    return render(request, "portal/admin_vacancies_form.html", {"form": form, "mode": "edit", "obj": obj})


@staff_member_required
@require_POST
def admin_vacancies_toggle(request, pk: int):
    obj = get_object_or_404(SchoolVacancy, pk=pk)
    obj.is_open = not obj.is_open
    obj.save(update_fields=["is_open"])
    messages.success(request, "تم تحديث حالة الشاغر.")
    return _redirect_admin_vacancies_list_with_filters(request)


@staff_member_required
@require_POST
def admin_vacancies_disable_all_view(request):
    selected_ids = [x for x in request.POST.getlist("selected_vacancies") if str(x).isdigit()]
    scope = (request.POST.get("scope") or "").strip()

    if scope == "selected":
        if not selected_ids:
            messages.warning(request, "لم يتم تحديد أي مدرسة/شاغر للإغلاق.")
            return _redirect_admin_vacancies_list_with_filters(request)

        updated = SchoolVacancy.objects.filter(id__in=selected_ids, is_open=True).update(is_open=False)
        messages.success(request, f"تم إغلاق المدارس/الشواغر المحددة بنجاح. العدد المتأثر: {updated}")
        return _redirect_admin_vacancies_list_with_filters(request)

    updated = SchoolVacancy.objects.filter(is_open=True).update(is_open=False)
    messages.success(request, f"تم تعطيل جميع المدارس/الشواغر بنجاح. العدد المتأثر: {updated}")
    return _redirect_admin_vacancies_list_with_filters(request)


@staff_member_required
@require_POST
def admin_vacancies_enable_all_view(request):
    selected_ids = [x for x in request.POST.getlist("selected_vacancies") if str(x).isdigit()]
    scope = (request.POST.get("scope") or "").strip()

    if scope == "selected":
        if not selected_ids:
            messages.warning(request, "لم يتم تحديد أي مدرسة/شاغر للفتح.")
            return _redirect_admin_vacancies_list_with_filters(request)

        updated = SchoolVacancy.objects.filter(id__in=selected_ids, is_open=False).update(is_open=True)
        messages.success(request, f"تم فتح المدارس/الشواغر المحددة بنجاح. العدد المتأثر: {updated}")
        return _redirect_admin_vacancies_list_with_filters(request)

    updated = SchoolVacancy.objects.filter(is_open=False).update(is_open=True)
    messages.success(request, f"تم تفعيل جميع المدارس/الشواغر بنجاح. العدد المتأثر: {updated}")
    return _redirect_admin_vacancies_list_with_filters(request)


@staff_member_required
@require_POST
def admin_vacancies_delete(request, pk: int):
    if not request.user.is_superuser:
        messages.error(request, "غير مصرح بالحذف النهائي. استخدم الإغلاق بدلًا من ذلك.")
        return _redirect_admin_vacancies_list_with_filters(request)

    obj = get_object_or_404(SchoolVacancy, pk=pk)

    if ApplicationPreference.objects.filter(vacancy=obj).exists():
        messages.error(request, "لا يمكن الحذف النهائي: يوجد رغبات مرتبطة بهذا الشاغر. استخدم (إغلاق) بدلًا من ذلك.")
        return _redirect_admin_vacancies_list_with_filters(request)

    obj.delete()
    messages.success(request, "تم حذف الشاغر نهائيًا.")
    return _redirect_admin_vacancies_list_with_filters(request)


# =========================================================
# Admin: Final Approvals Helpers
# =========================================================
def _final_approvals_filters_from_request(request):
    q = (request.GET.get("q") or "").strip()
    sector = (request.GET.get("sector") or "").strip()
    gender = (request.GET.get("gender") or "").strip()
    achieved_only = (request.GET.get("achieved_only") or "").strip()
    decision_type = (request.GET.get("decision_type") or "").strip()

    # توافق خلفي مع خيار "المتحققة فقط" القديم.
    if achieved_only == "1" and not decision_type:
        decision_type = "achieved"

    allowed_types = {"", "achieved", "pending_achieved", "documented_no_prefs"}
    if decision_type not in allowed_types:
        decision_type = ""

    return q, sector, gender, achieved_only, decision_type


def _final_approvals_base_qs(request):
    q, sector, gender, _achieved_only, _decision_type = _final_approvals_filters_from_request(request)

    qs = (
        Application.objects
        .select_related(
            "applicant",
            "achieved_pref__vacancy",
            "admin_decided_by",
            "achieved_by",
        )
        .prefetch_related("prefs", "prefs__vacancy")
        .filter(admin_decision="approved")
        .order_by("-admin_decided_at", "-id")
    )

    if sector:
        qs = qs.filter(applicant__sector__icontains=sector)

    if gender:
        qs = qs.filter(applicant__gender__icontains=gender)

    if q:
        qs = qs.filter(
            Q(applicant__full_name__icontains=q)
            | Q(applicant__national_id__icontains=q)
            | Q(applicant__sector__icontains=q)
            | Q(applicant__gender__icontains=q)
            | Q(achieved_pref__vacancy__school_name__icontains=q)
            | Q(achieved_pref__vacancy__stage__icontains=q)
            | Q(achieved_pref__vacancy__sector__icontains=q)
        )

    return qs


def _apply_final_approvals_decision_type_filter(qs, decision_type: str):
    """
    مخرجات القرار الإداري بعد اعتماد الإدارة:
    - achieved: معتمد وله رغبة متحققة.
    - pending_achieved: معتمد وله رغبات لكنه بانتظار تحديد رغبة نهائية.
    - documented_no_prefs: موثق الاستلام دون رغبات، وليس بانتظار رغبة.
    """
    decision_type = (decision_type or "").strip()

    if decision_type == "achieved":
        return qs.filter(achieved_pref__isnull=False)

    if decision_type == "documented_no_prefs":
        return qs.filter(status="submitted", prefs__isnull=True).distinct()

    if decision_type == "pending_achieved":
        return qs.filter(
            status="submitted",
            achieved_pref__isnull=True,
            prefs__isnull=False,
        ).distinct()

    return qs


def _final_approvals_qs(request):
    _q, _sector, _gender, _achieved_only, decision_type = _final_approvals_filters_from_request(request)
    qs = _final_approvals_base_qs(request)
    return _apply_final_approvals_decision_type_filter(qs, decision_type)


def _final_approval_output_type(app: Application) -> str:
    if _is_submitted_without_preferences(app):
        return "موثق الاستلام"
    if getattr(app, "achieved_pref_id", None):
        return "معتمد برغبة متحققة"
    if _is_submitted_with_preferences(app):
        return "معتمد بانتظار تحديد رغبة"
    return _admin_decision_display(app).get("label", "معتمد")


# =========================================================
# Admin: Final Approvals
# =========================================================
@staff_member_required
def admin_final_approvals_view(request):
    q, sector, gender, achieved_only, decision_type = _final_approvals_filters_from_request(request)
    qs = _final_approvals_qs(request)
    base_qs = _final_approvals_base_qs(request)

    page_obj = _paginate(request, qs, per_page=40)
    for app in page_obj.object_list:
        _enrich_admin_application(app)
        app.final_output_type = _final_approval_output_type(app)

    total = qs.count()
    total_outputs = base_qs.count()
    total_achieved = base_qs.filter(achieved_pref__isnull=False).count()
    total_documented_no_prefs = base_qs.filter(status="submitted", prefs__isnull=True).distinct().count()
    total_pending_achieved = base_qs.filter(
        status="submitted",
        achieved_pref__isnull=True,
        prefs__isnull=False,
    ).distinct().count()

    sectors = list(
        Applicant.objects
        .exclude(sector__isnull=True)
        .exclude(sector__exact="")
        .values_list("sector", flat=True)
        .distinct()
        .order_by("sector")
    )

    ctx = {
        "rows": page_obj,
        "q": q,
        "sector": sector,
        "gender": gender,
        "achieved_only": achieved_only,
        "decision_type": decision_type,
        "total": total,
        "total_outputs": total_outputs,
        "total_achieved": total_achieved,
        "total_documented_no_prefs": total_documented_no_prefs,
        "total_pending_achieved": total_pending_achieved,
        "sectors": sectors,
    }
    return render(request, "portal/admin_final_approvals.html", ctx)


@staff_member_required
def admin_final_approvals_print_view(request):
    q, sector, gender, achieved_only, decision_type = _final_approvals_filters_from_request(request)
    qs = _final_approvals_qs(request)
    base_qs = _final_approvals_base_qs(request)

    rows = list(qs[:5000])
    for app in rows:
        _enrich_admin_application(app)
        app.final_output_type = _final_approval_output_type(app)

    total = qs.count()
    total_outputs = base_qs.count()
    total_achieved = base_qs.filter(achieved_pref__isnull=False).count()
    total_documented_no_prefs = base_qs.filter(status="submitted", prefs__isnull=True).distinct().count()
    total_pending_achieved = base_qs.filter(
        status="submitted",
        achieved_pref__isnull=True,
        prefs__isnull=False,
    ).distinct().count()

    ctx = {
        "rows": rows,
        "q": q,
        "sector": sector,
        "gender": gender,
        "achieved_only": achieved_only,
        "decision_type": decision_type,
        "total": total,
        "total_outputs": total_outputs,
        "total_achieved": total_achieved,
        "total_documented_no_prefs": total_documented_no_prefs,
        "total_pending_achieved": total_pending_achieved,
        "now": timezone.localtime(),
        "print_mode": True,
    }
    return render(request, "portal/admin_final_approvals.html", ctx)


@staff_member_required
def admin_final_approvals_excel_view(request):
    qs = _final_approvals_qs(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Decision Outputs"

    headers = [
        "#",
        "رقم الطلب",
        "الاسم",
        "السجل المدني",
        "القطاع",
        "الجنس",
        "نوع المخرج الإداري",
        "قرار الإدارة",
        "يدخل مفاضلة الرغبات؟",
        "قابل للمعالجة الإدارية؟",
        "الرغبة المتحققة",
        "المدرسة النهائية",
        "مرحلة المدرسة",
        "قطاع المدرسة",
        "الملاحظة/الأثر الإداري",
        "تاريخ القرار/التوثيق",
        "قرر بواسطة",
        "تاريخ التحقق النهائي",
        "تحقق بواسطة",
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, app in enumerate(qs, start=1):
        prefs = list(app.prefs.select_related("vacancy").order_by("rank", "id"))
        _enrich_admin_application(app, prefs)

        no_prefs = bool(getattr(app, "is_no_preferences_path", False))
        vac = app.achieved_pref.vacancy if app.achieved_pref else None

        if no_prefs:
            achieved_text = "غير منطبق — لا توجد رغبات مسجلة"
            school_text = "لا توجد مدرسة محددة"
            stage_text = "—"
            school_sector_text = "—"
            effect_note = "موثق الاستلام دون رغبات؛ لا يدخل مفاضلة الرغبات، ولا يترتب عليه مطالبة بشاغر محدد، مع احتفاظ الإدارة بحق المعالجة وفق المصلحة التعليمية والاحتياج والضوابط المعتمدة."
        elif app.achieved_pref:
            achieved_text = f"رغبة {app.achieved_pref.rank}"
            school_text = getattr(vac, "school_name", "") if vac else ""
            stage_text = getattr(vac, "stage", "") if vac else ""
            school_sector_text = getattr(vac, "sector", "") if vac else ""
            effect_note = "معتمد برغبة متحققة وفق الضوابط ونتائج المفاضلة والاحتياج."
        else:
            achieved_text = "بانتظار تحديد رغبة نهائية"
            school_text = "—"
            stage_text = "—"
            school_sector_text = "—"
            effect_note = "معتمد إداريًا ولديه رغبات مسجلة، لكنه بانتظار تحديد الرغبة النهائية المتحققة."

        ws.append([
            i,
            app.id,
            getattr(app.applicant, "full_name", "") or "",
            getattr(app.applicant, "national_id", "") or "",
            getattr(app.applicant, "sector", "") or "",
            getattr(app.applicant, "gender", "") or "",
            _final_approval_output_type(app),
            getattr(app, "admin_decision_display", "") or _admin_decision_display(app, prefs).get("label", ""),
            getattr(app, "path_info", {}).get("competition_value", ""),
            getattr(app, "path_info", {}).get("admin_processing_value", ""),
            achieved_text,
            school_text,
            stage_text,
            school_sector_text,
            effect_note,
            _fmt_dt(app.admin_decided_at),
            getattr(app.admin_decided_by, "username", "") if app.admin_decided_by else "",
            _fmt_dt(app.achieved_at),
            getattr(app.achieved_by, "username", "") if app.achieved_by else "",
        ])

    widths = [6, 10, 28, 18, 18, 12, 24, 18, 18, 20, 24, 34, 16, 18, 60, 20, 16, 20, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"decision_outputs_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@staff_member_required
def admin_final_approvals_to_dashboard_view(request):
    """
    تحويل الفلاتر الحالية من صفحة مخرجات القرار إلى لوحة القرارات.
    """
    q, sector, gender, _achieved_only, _decision_type = _final_approvals_filters_from_request(request)

    params = {
        "decision": "approved",
    }
    if q:
        params["q"] = q
    if sector:
        params["sector"] = sector
    if gender:
        params["gender"] = gender

    url = f'{redirect("portal:admin_dashboard").url}?{urlencode(params)}'
    return redirect(url)


# =========================================================
# Admin: Vacancies Pressure Report
# =========================================================
def _vacancies_pressure_ctx(request):
    q = (request.GET.get("q") or "").strip()
    gender = (request.GET.get("gender") or "").strip()
    sector = (request.GET.get("sector") or "").strip()
    open_state = (request.GET.get("open") or "").strip()
    sort = (request.GET.get("sort") or "rank1").strip()
    top_raw = (request.GET.get("top") or "0").strip()

    try:
        top = int(top_raw) if top_raw else 0
    except Exception:
        top = 0

    achieved_sq = (
        Application.objects
        .filter(achieved_pref__vacancy_id=OuterRef("pk"))
        .values("achieved_pref__vacancy_id")
        .annotate(c=Count("id"))
        .values("c")[:1]
    )

    qs = (
        SchoolVacancy.objects
        .all()
        .annotate(
            interested_total=Count("applicationpreference", distinct=True),
            interested_rank1=Count(
                "applicationpreference",
                filter=Q(applicationpreference__rank=1),
                distinct=True,
            ),
            achieved_total=Coalesce(Subquery(achieved_sq, output_field=IntegerField()), Value(0)),
        )
    )

    if open_state == "open":
        qs = qs.filter(is_open=True)
    elif open_state == "closed":
        qs = qs.filter(is_open=False)

    if gender:
        qs = qs.filter(gender__icontains=gender)

    if sector:
        qs = qs.filter(sector__icontains=sector)

    if q:
        qs = qs.filter(
            Q(school_name__icontains=q)
            | Q(ministry_no__icontains=q)
            | Q(sector__icontains=q)
            | Q(manager_name__icontains=q)
        )

    if sort == "total":
        qs = qs.order_by("-interested_total", "-interested_rank1", "-achieved_total", "school_name")
    elif sort == "achieved":
        qs = qs.order_by("-achieved_total", "-interested_rank1", "-interested_total", "school_name")
    elif sort == "need":
        qs = qs.order_by("-deputy_need", "-interested_rank1", "-interested_total", "school_name")
    else:
        qs = qs.order_by("-interested_rank1", "-interested_total", "-achieved_total", "school_name")

    rows = list(qs[:top] if top and top > 0 else qs[:5000])

    total_schools = qs.count()
    sum_total = sum(int(getattr(x, "interested_total", 0) or 0) for x in rows)
    sum_rank1 = sum(int(getattr(x, "interested_rank1", 0) or 0) for x in rows)
    sum_achieved = sum(int(getattr(x, "achieved_total", 0) or 0) for x in rows)

    return {
        "rows": rows,
        "total_schools": total_schools,
        "sum_total": sum_total,
        "sum_rank1": sum_rank1,
        "sum_achieved": sum_achieved,
        "now": timezone.localtime(),
        "f": {
            "q": q,
            "sector": sector,
            "gender": gender,
            "open": open_state,
            "sort": sort,
            "top": top,
        },
    }


@staff_member_required
def admin_vacancies_pressure_report_view(request):
    return render(request, "portal/admin_vacancies_pressure_report.html", _vacancies_pressure_ctx(request))


@staff_member_required
def admin_vacancies_pressure_print_view(request):
    return render(request, "portal/admin_vacancies_pressure_print.html", _vacancies_pressure_ctx(request))


@staff_member_required
def admin_vacancies_pressure_csv_view(request):
    ctx = _vacancies_pressure_ctx(request)
    rows = ctx["rows"]

    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = 'attachment; filename="vacancies_pressure.csv"'
    resp.write("\ufeff")

    w = csv.writer(resp)
    w.writerow([
        "#",
        "المدرسة",
        "رقم الوزارة",
        "القطاع",
        "الجنس",
        "المرحلة",
        "الاحتياج",
        "الراغبون",
        "رغبة أولى",
        "ترشيحات نهائية",
        "الحالة",
    ])

    for i, v in enumerate(rows, start=1):
        w.writerow([
            i,
            v.school_name,
            v.ministry_no,
            v.sector,
            v.gender,
            v.stage,
            v.deputy_need,
            getattr(v, "interested_total", 0) or 0,
            getattr(v, "interested_rank1", 0) or 0,
            getattr(v, "achieved_total", 0) or 0,
            "مفتوح" if v.is_open else "مغلق",
        ])

    return resp


@staff_member_required
def admin_vacancies_pressure_excel_view(request):
    ctx = _vacancies_pressure_ctx(request)
    rows = ctx["rows"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Pressure"

    headers = [
        "#", "المدرسة", "رقم الوزارة", "القطاع", "الجنس", "المرحلة",
        "الاحتياج", "الراغبون", "رغبة أولى", "ترشيحات نهائية", "الحالة",
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, v in enumerate(rows, start=1):
        ws.append([
            i,
            v.school_name,
            v.ministry_no,
            v.sector,
            v.gender,
            v.stage,
            v.deputy_need,
            getattr(v, "interested_total", 0) or 0,
            getattr(v, "interested_rank1", 0) or 0,
            getattr(v, "achieved_total", 0) or 0,
            "مفتوح" if v.is_open else "مغلق",
        ])

    widths = [5, 42, 14, 20, 10, 14, 10, 12, 12, 14, 10]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"vacancies_pressure_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# =========================================================
# Admin: Dashboard + Reading
# =========================================================
@staff_member_required
def admin_dashboard_view(request):
    q, status, sector, gender, decision = _admin_filters_from_request(request)

    qs0 = _admin_base_qs()
    qs = _apply_admin_filters(qs0, q, status, sector, gender, decision)

    pref_rank1 = (
        ApplicationPreference.objects
        .filter(application=OuterRef("pk"), rank=1)
        .select_related("vacancy")
        .order_by("id")
    )

    qs = qs.annotate(
        first_school=Subquery(pref_rank1.values("vacancy__school_name")[:1], output_field=CharField()),
        first_stage=Subquery(pref_rank1.values("vacancy__stage")[:1], output_field=CharField()),
    ).annotate(
        prefs_count=Count("prefs", distinct=True),
        first_pref_text=Case(
            When(first_school__isnull=True, then=Value("-")),
            default=Concat(
                Coalesce("first_school", Value("")),
                Value(" — "),
                Coalesce("first_stage", Value("")),
                output_field=CharField(),
            ),
            output_field=CharField(),
        ),
    )

    qs = qs.annotate(
        achieved_text=Case(
            When(achieved_pref__isnull=True, then=Value("")),
            default=Concat(
                Value("رغبة "),
                Cast(Coalesce("achieved_pref__rank", Value(0)), output_field=CharField()),
                Value(" — "),
                Coalesce("achieved_pref__vacancy__school_name", Value("")),
                Value(" ("),
                Coalesce("achieved_pref__vacancy__stage", Value("")),
                Value(")"),
                output_field=CharField(),
            ),
            output_field=CharField(),
        )
    )



    pending_issues_sq = (
        ApplicantDataIssue.objects
        .filter(
            applicant_id=OuterRef("applicant_id"),
            status=ApplicantDataIssue.STATUS_PENDING,
        )
        .order_by()
        .values("applicant_id")
        .annotate(c=Count("id"))
        .values("c")[:1]
    )

    pending_blocking_issues_sq = (
        ApplicantDataIssue.objects
        .filter(
            applicant_id=OuterRef("applicant_id"),
            status=ApplicantDataIssue.STATUS_PENDING,
            is_blocking=True,
        )
        .order_by()
        .values("applicant_id")
        .annotate(c=Count("id"))
        .values("c")[:1]
    )

    qs = qs.annotate(
        pending_data_issues_count=Coalesce(
            Subquery(pending_issues_sq, output_field=IntegerField()),
            Value(0),
        ),
        pending_blocking_data_issues_count=Coalesce(
            Subquery(pending_blocking_issues_sq, output_field=IntegerField()),
            Value(0),
        ),
    )

    rows = list(qs[:500])
    for row in rows:
        _enrich_admin_application(row)

    total_apps = qs.count()

    total_prefs = (
        ApplicationPreference.objects
        .filter(application__in=qs.values("id"))
        .count()
    )

    unique_sectors = (
        qs.values("applicant__sector")
        .exclude(applicant__sector__isnull=True)
        .exclude(applicant__sector__exact="")
        .distinct()
        .count()
    )

    status_counts = list(qs.values("status").annotate(c=Count("id")).order_by("-c"))
    for it in status_counts:
        it["label"] = (it.get("status") or "-")
        it["pct"] = _pct(int(it.get("c") or 0), total_apps)

    decision_counts = list(qs.values("admin_decision").annotate(c=Count("id")).order_by("-c"))
    for it in decision_counts:
        raw = (it.get("admin_decision") or "").strip() or "pending"
        it["label"] = raw
        it["pct"] = _pct(int(it.get("c") or 0), total_apps)

    count_submitted = qs.filter(status="submitted").count()
    count_draft = qs.filter(status="draft").count()
    nominated_count = qs.filter(achieved_pref__isnull=False).count()

    # مؤشرات ضابط الدخول والإرسال:
    # - دخل ولم يؤكد: تم إنشاء Application عند تسجيل الدخول، لكن لم يضغط تأكيد البيانات.
    # - أكد ولم يرسل: أكد بياناته وبقي الطلب Draft.
    # - مرسل بلا رغبات: ضغط إرسال الطلب بدون اختيار أي رغبة.
    # - مرسل برغبات: ضغط إرسال الطلب ومعه رغبة واحدة فأكثر.
    count_entered_not_confirmed = qs.filter(confirmed_at__isnull=True).count()
    count_confirmed_not_submitted = qs.filter(
        confirmed_at__isnull=False,
        status="draft",
    ).count()
    count_submitted_without_prefs = qs.filter(
        status="submitted",
        prefs__isnull=True,
    ).distinct().count()
    count_submitted_with_prefs = qs.filter(
        status="submitted",
        prefs__isnull=False,
    ).distinct().count()

    # الطلبات المشروطة بمراجعة البيانات:
    # المتقدم يستطيع إرسال رغباته حفظًا لحقه، لكن لا يدخل الطلب المفاضلة
    # ولا يعتمد إداريًا حتى تتم مراجعة طلب تعديل البيانات المؤثر.
    conditional_review_applicant_ids = (
        ApplicantDataIssue.objects
        .filter(
            status=ApplicantDataIssue.STATUS_PENDING,
            is_blocking=True,
        )
        .values("applicant_id")
    )

    count_conditional_data_review = qs.filter(
        status="submitted",
        applicant_id__in=conditional_review_applicant_ids,
    ).distinct().count()

    # قابل للمفاضلة: مرسل، لديه رغبات، ولم يصدر عليه قرار إداري بعد،
    # وليس معلقًا على مراجعة بيانات مؤثرة.
    count_competition_ready = qs.filter(
        status="submitted",
        prefs__isnull=False,
    ).filter(
        Q(admin_decision__isnull=True) | Q(admin_decision__exact="")
    ).exclude(
        applicant_id__in=conditional_review_applicant_ids,
    ).distinct().count()

    sectors = list(
        Applicant.objects
        .exclude(sector__isnull=True)
        .exclude(sector__exact="")
        .values_list("sector", flat=True)
        .distinct()
        .order_by("sector")
    )


    pending_data_issues_count = ApplicantDataIssue.objects.filter(
        status=ApplicantDataIssue.STATUS_PENDING,
    ).count()

    pending_blocking_data_issues_count = ApplicantDataIssue.objects.filter(
        status=ApplicantDataIssue.STATUS_PENDING,
        is_blocking=True,
    ).count()

    data_issues_total_count = ApplicantDataIssue.objects.count()

    protected_pending_data_issues_count = ApplicantDataIssue.objects.filter(
        status=ApplicantDataIssue.STATUS_PENDING,
        is_blocking=True,
        protects_followup_right=True,
    ).count()
    active_followup_windows_count = ApplicantDataIssue.objects.filter(
        protects_followup_right=True,
        followup_window_expires_at__gte=timezone.now(),
        status__in=[
            ApplicantDataIssue.STATUS_ALLOWED,
            ApplicantDataIssue.STATUS_CORRECTED,
            ApplicantDataIssue.STATUS_REJECTED,
        ],
    ).count()

    portal_window = PortalWindow.get()

    current_query = request.GET.urlencode()
    query_suffix = f"?{current_query}" if current_query else ""

    ctx = {
        "rows": rows,
        "total_apps": total_apps,
        "total_prefs": total_prefs,
        "unique_sectors": unique_sectors,
        "status_counts": status_counts,
        "decision_counts": decision_counts,
        "count_submitted": count_submitted,
        "count_draft": count_draft,
        "nominated_count": nominated_count,
        "count_entered_not_confirmed": count_entered_not_confirmed,
        "count_confirmed_not_submitted": count_confirmed_not_submitted,
        "count_submitted_without_prefs": count_submitted_without_prefs,
        "count_submitted_with_prefs": count_submitted_with_prefs,
        "count_competition_ready": count_competition_ready,
        "count_conditional_data_review": count_conditional_data_review,
        "pending_data_issues_count": pending_data_issues_count,
        "pending_blocking_data_issues_count": pending_blocking_data_issues_count,
        "data_issues_total_count": data_issues_total_count,
        "protected_pending_data_issues_count": protected_pending_data_issues_count,
        "active_followup_windows_count": active_followup_windows_count,
        "f_q": q,
        "f_status": status,
        "f_sector": sector,
        "f_gender": gender,
        "f_decision": decision,
        "sectors": sectors,
        "portal_window": portal_window,
        "current_query": current_query,
        "current_path": request.get_full_path(),

        # روابط قائمة المزيد
        "url_general_excel": f'{redirect("portal:admin_export_excel").url}{query_suffix}',
        "url_general_print": f'{redirect("portal:admin_report_print").url}{query_suffix}',
        "url_school_demand_report": f'{redirect("portal:admin_vacancies_pressure").url}{query_suffix}',
        "url_candidates_csv": f'{redirect("portal:admin_nominations_csv").url}{query_suffix}',
        "url_candidates_excel": f'{redirect("portal:admin_nominations_excel").url}{query_suffix}',
        "url_candidates_print": f'{redirect("portal:admin_nominations_print").url}{query_suffix}',
        "url_candidates_report": f'{redirect("portal:admin_nominations_report").url}{query_suffix}',
    }
    return render(request, "portal/admin_dashboard.html", ctx)


@staff_member_required
def admin_application_detail_view(request, app_id: int):
    app = get_object_or_404(
        Application.objects.select_related(
            "applicant",
            "achieved_by",
            "admin_decided_by",
            "achieved_pref__vacancy",
        ),
        id=app_id,
    )

    prefs = list(
        ApplicationPreference.objects
        .filter(application=app)
        .select_related("vacancy")
        .order_by("rank")
    )

    back_url = (request.GET.get("back") or "").strip()
    if not back_url:
        back_url = redirect("portal:admin_dashboard").url

    _enrich_admin_application(app, prefs)
    path_info = app.path_info
    decision_info = app.decision_info
    proof = _submission_proof_context(app, prefs)

    return render(
        request,
        "portal/admin_application_detail.html",
        {
            "app": app,
            "a": app.applicant,
            "prefs": prefs,
            "back_url": back_url,
            "path_info": path_info,
            "decision_info": decision_info,
            "proof": proof,
            "has_admin_decision": bool((getattr(app, "admin_decision", "") or "").strip()),
            "prefs_count": len(prefs),
        },
    )


@staff_member_required
def admin_application_print_view(request, app_id: int):
    application = get_object_or_404(
        Application.objects.select_related(
            "applicant",
            "achieved_pref__vacancy",
            "achieved_by",
            "admin_decided_by",
        ),
        id=app_id,
    )

    prefs = list(
        ApplicationPreference.objects
        .filter(application=application)
        .select_related("vacancy")
        .order_by("rank")
    )
    selected_pref = application.achieved_pref if getattr(application, "achieved_pref_id", None) else None
    _enrich_admin_application(application, prefs)
    path_info = application.path_info
    decision_info = application.decision_info
    proof = _submission_proof_context(application, prefs)

    return render(
        request,
        "portal/admin_application_print.html",
        {
            "application": application,
            "app": application,
            "a": application.applicant,
            "prefs": prefs,
            "selected_pref": selected_pref,
            "path_info": path_info,
            "decision_info": decision_info,
            "proof": proof,
        },
    )


@staff_member_required
def admin_application_detail_json_view(request, app_id: int):
    app = get_object_or_404(
        Application.objects.select_related("applicant", "achieved_pref__vacancy"),
        id=app_id,
    )

    prefs = list(
        ApplicationPreference.objects
        .filter(application=app)
        .select_related("vacancy")
        .order_by("rank")
    )

    achieved = None
    if getattr(app, "achieved_pref_id", None) and getattr(app, "achieved_pref", None):
        ap = app.achieved_pref
        if ap and getattr(ap, "vacancy", None):
            achieved = {
                "pref_id": ap.id,
                "rank": ap.rank,
                "label": f"{ap.vacancy.school_name} — {ap.vacancy.stage}",
            }

    data = {
        "id": app.id,
        "status": app.status,
        "submitted_at": _fmt_dt(app.submitted_at),
        "admin_decision": getattr(app, "admin_decision", "") or "",
        "admin_note": getattr(app, "admin_note", "") or "",
        "admin_decided_at": _fmt_dt(getattr(app, "admin_decided_at", None)),
        "achieved": achieved,
        "applicant": {
            "national_id": app.applicant.national_id,
            "full_name": app.applicant.full_name,
            "sector": app.applicant.sector,
            "gender": app.applicant.gender,
        },
        "decision_display": _admin_decision_display(app, prefs),
        "path_info": _application_path_info(app, prefs),
        "proof": _submission_proof_context(app, prefs),
        "prefs": [{"id": p.id, "rank": p.rank, "label": f"{p.vacancy.school_name} — {p.vacancy.stage}"} for p in prefs],
    }
    return JsonResponse(data, json_dumps_params={"ensure_ascii": False})


# =========================================================
# Admin: Decision Actions
# =========================================================
@staff_member_required
@require_POST
def admin_decide_approve_view(request, app_id: int):
    app = get_object_or_404(Application, id=app_id)
    note = (request.POST.get("note") or "").strip()
    back_url = (request.GET.get("back") or "").strip()

    conditional_issue = _application_conditional_data_issue(app)
    if conditional_issue:
        msg = "لا يمكن اعتماد الطلب قبل مراجعة طلب تعديل البيانات المؤثر. راجع طلبات التعديل أولًا."
        if _is_ajax(request):
            return JsonResponse({"ok": False, "error": msg}, status=400, json_dumps_params={"ensure_ascii": False})
        messages.error(request, msg)
        return _redirect_admin_app_detail_with_back(app.id, back_url)

    is_no_prefs = _is_submitted_without_preferences(app)
    if is_no_prefs and not note:
        note = NO_PREFERENCES_ADMIN_DECISION_NOTE

    _set_admin_decision(app, request.user, "approved", note)

    if _is_ajax(request):
        return JsonResponse(
            {
                "ok": True,
                "id": app.id,
                "admin_decision": "approved",
                "no_preferences_path": is_no_prefs,
                "message": "تم توثيق استلام الطلب دون رغبات." if is_no_prefs else "تم اعتماد الطلب.",
            },
            json_dumps_params={"ensure_ascii": False},
        )

    if is_no_prefs:
        messages.success(request, f"تم توثيق استلام الطلب دون رغبات #{app.id}")
    else:
        messages.success(request, f"تم اعتماد الطلب #{app.id}")
    return _redirect_admin_app_detail_with_back(app.id, back_url)


@staff_member_required
@require_POST
def admin_decide_reject_view(request, app_id: int):
    app = get_object_or_404(Application, id=app_id)
    note = (request.POST.get("note") or "").strip()
    back_url = (request.GET.get("back") or "").strip()
    if not note:
        if _is_ajax(request):
            return JsonResponse({"ok": False, "error": "فضلاً اكتب سبب الرفض."}, status=400, json_dumps_params={"ensure_ascii": False})
        messages.error(request, "فضلاً اكتب سبب الرفض.")
        return _redirect_admin_app_detail_with_back(app.id, back_url)

    _set_admin_decision(app, request.user, "rejected", note)

    if _is_ajax(request):
        return JsonResponse({"ok": True, "id": app.id, "admin_decision": "rejected"}, json_dumps_params={"ensure_ascii": False})

    messages.success(request, f"تم رفض الطلب #{app.id}")
    return _redirect_admin_app_detail_with_back(app.id, back_url)


@staff_member_required
@require_POST
def admin_decide_unlock_view(request, app_id: int):
    app = get_object_or_404(Application, id=app_id)
    note = (request.POST.get("note") or "").strip()
    back_url = (request.GET.get("back") or "").strip()
    if not note:
        if _is_ajax(request):
            return JsonResponse({"ok": False, "error": "فضلاً اكتب سبب الإرجاع للتعديل."}, status=400, json_dumps_params={"ensure_ascii": False})
        messages.error(request, "فضلاً اكتب سبب الإرجاع للتعديل.")
        return _redirect_admin_app_detail_with_back(app.id, back_url)

    app.locked = False
    app.status = "draft"
    app.save(update_fields=["locked", "status"])

    _set_admin_decision(app, request.user, "returned", note)

    if _is_ajax(request):
        return JsonResponse({"ok": True, "id": app.id, "admin_decision": "returned", "status": "draft"}, json_dumps_params={"ensure_ascii": False})

    messages.success(request, f"تم فتح التعديل للطلب #{app.id}")
    return _redirect_admin_app_detail_with_back(app.id, back_url)


# =========================================================
# Admin: Undo Last Decision
# =========================================================
@staff_member_required
@require_POST
def admin_undo_view(request, app_id: int):
    if not _is_ajax(request):
        return JsonResponse(
            {"ok": False, "error": "AJAX فقط."},
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    app = get_object_or_404(Application, id=app_id)

    prev_status = (request.POST.get("prev_status") or "").strip()
    if prev_status not in {"draft", "submitted", "returned", "approved", "rejected"}:
        prev_status = "submitted"

    with transaction.atomic():
        app.status = prev_status
        app.locked = (prev_status == "submitted")

        app.admin_decision = ""
        app.admin_note = ""
        app.admin_decided_at = None
        app.admin_decided_by = None

        app.save(update_fields=[
            "status",
            "locked",
            "admin_decision",
            "admin_note",
            "admin_decided_at",
            "admin_decided_by",
        ])

    return JsonResponse(
        {
            "ok": True,
            "id": app.id,
            "status": app.status,
            "admin_decision": "",
        },
        json_dumps_params={"ensure_ascii": False},
    )


# =========================================================
# Admin: Bulk Decision
# =========================================================
@staff_member_required
@require_POST
def admin_decide_bulk_view(request):
    ids = request.POST.getlist("ids")
    action = (request.POST.get("action") or "").strip()
    note = (request.POST.get("note") or "").strip()

    clean_ids: list[int] = []
    for x in ids:
        try:
            clean_ids.append(int(x))
        except Exception:
            continue

    if not clean_ids:
        return JsonResponse({"ok": False, "error": "لا توجد طلبات محددة."}, status=400, json_dumps_params={"ensure_ascii": False})

    if action not in ("approve", "reject", "unlock"):
        return JsonResponse({"ok": False, "error": "إجراء غير صحيح."}, status=400, json_dumps_params={"ensure_ascii": False})

    if action in ("reject", "unlock") and not note:
        return JsonResponse({"ok": False, "error": "الملاحظة مطلوبة للرفض أو الإرجاع."}, status=400, json_dumps_params={"ensure_ascii": False})

    qs = Application.objects.filter(id__in=clean_ids)

    skipped_conditional = 0
    with transaction.atomic():
        updated = 0

        if action == "unlock":
            qs.update(locked=False, status="draft")
            for app in qs.select_for_update():
                _set_admin_decision(app, request.user, "returned", note)
                updated += 1

        elif action == "approve":
            for app in qs.select_for_update():
                if _application_is_conditional_data_review(app):
                    skipped_conditional += 1
                    continue
                note_for_app = note
                if _is_submitted_without_preferences(app) and not note_for_app:
                    note_for_app = NO_PREFERENCES_ADMIN_DECISION_NOTE
                _set_admin_decision(app, request.user, "approved", note_for_app)
                updated += 1

        else:
            for app in qs.select_for_update():
                _set_admin_decision(app, request.user, "rejected", note)
                updated += 1

    payload = {"ok": True, "updated": updated}
    if skipped_conditional:
        payload["skipped_conditional"] = skipped_conditional
        payload["warning"] = f"تم تجاوز {skipped_conditional} طلبًا لأنها مرسلة مشروطة بمراجعة البيانات."
    return JsonResponse(payload, json_dumps_params={"ensure_ascii": False})


# =========================================================
# Admin: Reports
# =========================================================
@staff_member_required
def admin_report_print_view(request):
    q, status, sector, gender, decision = _admin_filters_from_request(request)

    qs0 = (
        Application.objects
        .select_related("applicant")
        .annotate(prefs_count=Count("prefs", distinct=True))
        .order_by("-submitted_at", "-id")
    )
    qs = _apply_admin_filters(qs0, q, status, sector, gender, decision)
    rows = list(qs[:5000])

    ctx = {
        "rows": rows,
        "total": len(rows),
        "now": timezone.localtime(),
        "f": {"q": q, "status": status, "sector": sector, "gender": gender, "decision": decision},
    }
    return render(request, "portal/admin_report_print.html", ctx)


@staff_member_required
def admin_report_csv_visible_view(request):
    q, status, sector, gender, decision = _admin_filters_from_request(request)

    ids = (request.GET.get("ids") or "").strip()
    id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]

    qs0 = (
        Application.objects
        .select_related("applicant")
        .annotate(prefs_count=Count("prefs", distinct=True))
        .order_by("-submitted_at", "-id")
    )
    qs = _apply_admin_filters(qs0, q, status, sector, gender, decision)
    if id_list:
        qs = qs.filter(id__in=id_list)

    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = 'attachment; filename="decision_visible.csv"'
    resp.write("\ufeff")

    w = csv.writer(resp)
    w.writerow([
        "رقم الطلب", "الاسم", "السجل المدني", "القطاع", "الجنس",
        "الحالة", "عدد الرغبات", "قرار الإدارة", "ملاحظة الإدارة", "تاريخ التقديم",
    ])

    for app in qs:
        p = app.applicant
        w.writerow([
            app.id,
            getattr(p, "full_name", "") or "",
            getattr(p, "national_id", "") or "",
            getattr(p, "sector", "") or "",
            getattr(p, "gender", "") or "",
            getattr(app, "status", "") or "",
            getattr(app, "prefs_count", 0) or 0,
            getattr(app, "admin_decision", "") or "",
            getattr(app, "admin_note", "") or "",
            _fmt_dt(getattr(app, "submitted_at", None)),
        ])

    return resp


# =========================================================
# Admin: Set Achieved
# =========================================================
@staff_member_required
@require_POST
@transaction.atomic
def admin_set_achieved_view(request, app_id: int):
    app = get_object_or_404(
        Application.objects.select_related("applicant", "achieved_pref__vacancy"),
        id=app_id,
    )
    back_url = (request.GET.get("back") or "").strip()
    pref_id_raw = (request.POST.get("achieved_pref_id") or "").strip()

    old_vacancy = None
    if app.achieved_pref_id and app.achieved_pref and getattr(app.achieved_pref, "vacancy", None):
        old_vacancy = app.achieved_pref.vacancy

    if not pref_id_raw:
        if old_vacancy and old_vacancy.reserved_application_id == app.id:
            old_vacancy.reserved_application = None
            old_vacancy.reserved_at = None
            old_vacancy.save(update_fields=["reserved_application", "reserved_at"])

        app.achieved_pref = None
        app.achieved_at = None
        app.achieved_by = None
        app.save(update_fields=["achieved_pref", "achieved_at", "achieved_by"])

        messages.success(request, "تم إلغاء تحديد الرغبة المتحققة.")
        return _redirect_admin_app_detail_with_back(app.id, back_url)

    conditional_issue = _application_conditional_data_issue(app)
    if conditional_issue:
        messages.error(
            request,
            "لا يمكن تحديد رغبة متحققة أو إدخال الطلب في الترشيح قبل مراجعة طلب تعديل البيانات المؤثر."
        )
        return _redirect_admin_app_detail_with_back(app.id, back_url)

    try:
        pref_id = int(pref_id_raw)
    except ValueError:
        messages.error(request, "قيمة غير صحيحة.")
        return _redirect_admin_app_detail_with_back(app.id, back_url)

    pref = (
        ApplicationPreference.objects
        .filter(id=pref_id, application=app)
        .select_related("vacancy")
        .first()
    )
    if not pref:
        messages.error(request, "الرغبة المحددة غير تابعة لهذا الطلب.")
        return _redirect_admin_app_detail_with_back(app.id, back_url)

    vacancy = pref.vacancy

    if vacancy.reserved_application_id and vacancy.reserved_application_id != app.id:
        messages.error(request, "هذا الشاغر محجوز بالفعل لطلب آخر.")
        return _redirect_admin_app_detail_with_back(app.id, back_url)

    if (app.admin_decision or "").strip() != "approved":
        app.admin_decision = "approved"
        if not app.admin_decided_at:
            app.admin_decided_at = timezone.now()
        if not app.admin_decided_by_id:
            app.admin_decided_by = request.user
        app.save(update_fields=["admin_decision", "admin_decided_at", "admin_decided_by"])

    if old_vacancy and old_vacancy.id != vacancy.id and old_vacancy.reserved_application_id == app.id:
        old_vacancy.reserved_application = None
        old_vacancy.reserved_at = None
        old_vacancy.save(update_fields=["reserved_application", "reserved_at"])

    app.achieved_pref = pref
    app.achieved_at = timezone.now()
    app.achieved_by = request.user
    app.save(update_fields=["achieved_pref", "achieved_at", "achieved_by"])

    # عند تحديد أي مرشح على المدرسة يدويًا، يُحجز الشاغر ويُغلق تلقائيًا.
    # لا يُعاد فتحه لاحقًا إلا يدويًا من شاشة إدارة الشواغر.
    vacancy.reserved_application = app
    vacancy.reserved_at = timezone.now()
    vacancy.is_open = False
    vacancy.save(update_fields=["reserved_application", "reserved_at", "is_open"])

    messages.success(request, f"تم تحديد الرغبة المتحققة: رغبة #{pref.rank}")
    return _redirect_admin_app_detail_with_back(app.id, back_url)


# =========================================================
# Admin: Nominations
# =========================================================
@staff_member_required
def admin_nominations_report_view(request):
    qs = _nominations_qs(request)

    total = qs.count()
    by_sector = list(qs.values("applicant__sector").annotate(c=Count("id")).order_by("-c")[:12])
    for it in by_sector:
        it["label"] = (it.get("applicant__sector") or "-")
        it["pct"] = _pct(int(it.get("c") or 0), total)

    ctx = {
        "rows": list(qs[:5000]),
        "total": total,
        "by_sector": by_sector,
        "q": request.GET.get("q", ""),
        "sector": request.GET.get("sector", ""),
        "gender": request.GET.get("gender", ""),
        "school": request.GET.get("school", ""),
        "from_date": request.GET.get("from_date", ""),
        "to_date": request.GET.get("to_date", ""),
    }
    return render(request, "portal/admin_nominations_report.html", ctx)


@staff_member_required
def admin_nominations_print_view(request):
    qs = _nominations_qs(request)
    ctx = {
        "rows": qs,
        "total": qs.count(),
        "now": timezone.localtime(),
        "q": request.GET.get("q", ""),
        "sector": request.GET.get("sector", ""),
        "gender": request.GET.get("gender", ""),
        "school": request.GET.get("school", ""),
        "from_date": request.GET.get("from_date", ""),
        "to_date": request.GET.get("to_date", ""),
    }
    return render(request, "portal/admin_nominations_print.html", ctx)


@staff_member_required
def admin_nominations_csv_view(request):
    qs = _nominations_qs(request)

    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = 'attachment; filename="nominations.csv"'
    resp.write("\ufeff")

    w = csv.writer(resp)
    w.writerow([
        "#",
        "رقم الطلب",
        "الاسم",
        "السجل",
        "قطاع المتقدم",
        "جنس المتقدم",
        "الرغبة المتحققة",
        "مدرسة الترشيح",
        "رقم الوزارة",
        "مرحلة المدرسة",
        "قطاع المدرسة",
        "جنس المدرسة",
        "تاريخ الترشيح",
        "مرشح بواسطة",
        "قرار الإدارة",
        "ملاحظة الإدارة",
    ])

    for i, app in enumerate(qs, start=1):
        a = app.applicant
        vac = app.achieved_pref.vacancy if app.achieved_pref else None
        w.writerow([
            i,
            app.id,
            getattr(a, "full_name", "") or "",
            getattr(a, "national_id", "") or "",
            getattr(a, "sector", "") or "",
            getattr(a, "gender", "") or "",
            getattr(app.achieved_pref, "rank", "") if app.achieved_pref else "",
            getattr(vac, "school_name", "") if vac else "",
            getattr(vac, "ministry_no", "") if vac else "",
            getattr(vac, "stage", "") if vac else "",
            getattr(vac, "sector", "") if vac else "",
            getattr(vac, "gender", "") if vac else "",
            _fmt_dt(app.achieved_at),
            getattr(app.achieved_by, "username", "") if app.achieved_by else "",
            (app.admin_decision or "").strip(),
            (app.admin_note or "").strip(),
        ])

    return resp


@staff_member_required
def admin_nominations_excel_view(request):
    qs = _nominations_qs(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Nominations"

    headers = [
        "#",
        "رقم الطلب",
        "الاسم",
        "السجل",
        "قطاع المتقدم",
        "جنس المتقدم",
        "الرغبة المتحققة",
        "مدرسة الترشيح",
        "رقم الوزارة",
        "مرحلة المدرسة",
        "قطاع المدرسة",
        "جنس المدرسة",
        "تاريخ الترشيح",
        "مرشح بواسطة",
        "قرار الإدارة",
        "ملاحظة الإدارة",
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, app in enumerate(qs, start=1):
        a = app.applicant
        vac = app.achieved_pref.vacancy if app.achieved_pref else None

        ws.append([
            i,
            app.id,
            getattr(a, "full_name", "") or "",
            getattr(a, "national_id", "") or "",
            getattr(a, "sector", "") or "",
            getattr(a, "gender", "") or "",
            getattr(app.achieved_pref, "rank", "") if app.achieved_pref else "",
            getattr(vac, "school_name", "") if vac else "",
            getattr(vac, "ministry_no", "") if vac else "",
            getattr(vac, "stage", "") if vac else "",
            getattr(vac, "sector", "") if vac else "",
            getattr(vac, "gender", "") if vac else "",
            _fmt_dt(app.achieved_at),
            getattr(app.achieved_by, "username", "") if app.achieved_by else "",
            (app.admin_decision or "").strip(),
            (app.admin_note or "").strip(),
        ])

    widths = [5, 10, 28, 16, 18, 12, 14, 32, 14, 16, 18, 12, 18, 14, 14, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"nominations_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# =========================================================
# Admin: Export Excel
# =========================================================
@staff_member_required
def admin_export_excel_view(request):
    q, status, sector, gender, decision = _admin_filters_from_request(request)

    qs0 = (
        Application.objects
        .select_related("applicant", "achieved_pref__vacancy")
        .prefetch_related("prefs", "prefs__vacancy")
        .order_by("-submitted_at", "-id")
    )
    qs = _apply_admin_filters(qs0, q, status, sector, gender, decision)

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"

    def yes_no(value) -> str:
        return "نعم" if bool(value) else "لا"

    def decision_label(app: Application, prefs: list[ApplicationPreference]) -> str:
        return _admin_decision_display(app, prefs).get("label", "قيد المعالجة")

    def submission_type(app: Application, prefs: list[ApplicationPreference]) -> str:
        if app.status == "submitted":
            return "مرسل برغبات" if prefs else "مرسل دون رغبات"
        if getattr(app, "confirmed_at", None):
            return "أكد البيانات ولم يرسل"
        return "دخل ولم يؤكد البيانات"

    def administrative_status(app: Application, prefs: list[ApplicationPreference]) -> str:
        if app.status == "submitted" and prefs:
            return "مستلم للمعالجة / مقفل للتعديل"
        if app.status == "submitted" and not prefs:
            return "مستلم دون رغبات / مقفل للتعديل"
        if getattr(app, "locked", False):
            return "مقفل إداريًا كطلب غير مكتمل"
        return "غير مكتمل"

    def competition_status(app: Application, prefs: list[ApplicationPreference]) -> tuple[str, str]:
        if app.status != "submitted":
            return "لا", "لم يكتمل الإرسال النهائي"
        if not prefs:
            return "لا", "لم يسجل رغبات؛ لا يدخل في مفاضلة الرغبات ولا يترتب عليه مطالبة بشاغر محدد"
        return "نعم", PREFERENCES_COMPETITION_NOTE

    def administrative_processing_status(app: Application, prefs: list[ApplicationPreference]) -> tuple[str, str]:
        if app.status != "submitted":
            return "لا", "طلب غير مكتمل؛ لا يعالج كطلب مرسل إلا بعد الإرسال النهائي"
        if not prefs:
            return "نعم", "قابل للمعالجة الإدارية عند الحاجة وفق المصلحة التعليمية والاحتياج والضوابط، دون مطالبة بشاغر محدد"
        return "نعم", "قابل للمعالجة ضمن إجراءات المفاضلة والاعتماد النهائي وفق الضوابط والاحتياج"

    def achieved_label(app: Application, prefs: list[ApplicationPreference]) -> str:
        if getattr(app, "achieved_pref", None) and getattr(app.achieved_pref, "vacancy", None):
            return f"الرغبة {app.achieved_pref.rank} — {app.achieved_pref.vacancy.school_name}"
        if app.status == "submitted" and not prefs:
            return "غير منطبق — لا توجد رغبات مسجلة"
        if app.status == "submitted" and prefs:
            return "لم تتحقق رغبة حتى تاريخه"
        return "غير منطبق — طلب غير مكتمل"

    def prefs_count_at_submission(app: Application, prefs: list[ApplicationPreference]) -> int:
        saved_count = getattr(app, "submitted_prefs_count", None)
        if saved_count is not None:
            try:
                return int(saved_count)
            except Exception:
                pass
        return len(prefs)

    headers = [
        "ID",
        "رقم الهوية",
        "الاسم",
        "القطاع",
        "الجنس",
        "حالة النظام",
        "نوع الإرسال",
        "عدد الرغبات",
        "يدخل مفاضلة الرغبات؟",
        "سبب عدم الدخول / ملاحظة المفاضلة",
        "قابل للمعالجة الإدارية؟",
        "نطاق المعالجة الإدارية",
        "حالة الطلب إداريًا",
        "تاريخ الإرسال",
        "إقرار سياسة الرغبات",
        "وقت إقرار سياسة الرغبات",
        "إقرار دون رغبات",
        "وقت إقرار دون رغبات",
        "قرار الإدارة",
        "ملاحظة الإدارة",
        "تاريخ قرار الإدارة",
        "الرغبة المتحققة",
    ]
    for i in range(1, 11):
        headers.append(f"Pref {i}")
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for app in qs:
        prefs = sorted(list(app.prefs.all()), key=lambda p: p.rank)
        entered_competition, competition_reason = competition_status(app, prefs)
        administrative_processing, administrative_processing_note = administrative_processing_status(app, prefs)

        pref_names: list[str] = []
        if prefs:
            for p in prefs[:10]:
                v = p.vacancy
                pref_names.append(f"{p.rank}. {v.school_name} ({v.stage})")
        else:
            # لا نترك خانة الرغبة الأولى فارغة؛ لأن الفراغ في التقرير الإداري قابل للتأويل.
            if app.status == "submitted":
                pref_names.append("لم يتم اختيار رغبات")
            else:
                pref_names.append("لا توجد رغبات مسجلة")

        while len(pref_names) < 10:
            pref_names.append("—")

        row = [
            app.id,
            app.applicant.national_id,
            app.applicant.full_name,
            app.applicant.sector,
            app.applicant.gender,
            app.status,
            submission_type(app, prefs),
            prefs_count_at_submission(app, prefs),
            entered_competition,
            competition_reason,
            administrative_processing,
            administrative_processing_note,
            administrative_status(app, prefs),
            _fmt_dt(app.submitted_at),
            yes_no(getattr(app, "preferences_acknowledged", False)),
            _fmt_dt(getattr(app, "preferences_ack_at", None)),
            yes_no(getattr(app, "no_preferences_acknowledged", False)),
            _fmt_dt(getattr(app, "no_preferences_ack_at", None)),
            decision_label(app, prefs),
            (getattr(app, "admin_note", "") or "").strip(),
            _fmt_dt(getattr(app, "admin_decided_at", None)),
            achieved_label(app, prefs),
        ] + pref_names

        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    for col in range(1, ws.max_column + 1):
        max_len = 10
        for rowi in range(1, ws.max_row + 1):
            value = ws.cell(row=rowi, column=col).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 55)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"applications_export_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp



# =========================================================
# Admin: Import Excel
# =========================================================
def _activate_new_only_phase(force: bool = False) -> bool:
    win = PortalWindow.get()
    current_phase = _normalize_portal_phase(getattr(win, "phase", "closed"))

    if not force and current_phase in {"all", "official_only", "new_only"}:
        return False

    win.phase = "new_only"
    win.save(update_fields=["phase"])
    return True


def _handle_imported_applicants(uploaded_file, mode="sync"):
    path = _save_uploaded_file(uploaded_file, "applicants")
    batch, res = import_applicants_xlsx(path, mode=mode)
    portal_phase_switched = _activate_new_only_phase()
    return {
        "batch": batch.id,
        "created": res.created,
        "updated": res.updated,
        "skipped": res.skipped,
        "mode": mode,
        "portal_phase_switched": portal_phase_switched,
    }


def _handle_imported_schools(uploaded_file, mode="sync"):
    path = _save_uploaded_file(uploaded_file, "schools")
    batch, res = import_schools_xlsx(path, mode=mode)
    return {
        "batch": batch.id,
        "created": res.created,
        "updated": res.updated,
        "skipped": res.skipped,
        "mode": mode,
    }


@staff_member_required
@require_http_methods(["GET", "POST"])
def admin_import_view(request):
    form = ImportExcelForm(request.POST or None, request.FILES or None)
    result = {}

    if request.method == "POST" and form.is_valid():
        applicants_file = form.cleaned_data.get("applicants_file")
        schools_file = form.cleaned_data.get("schools_file")
        import_mode = (form.cleaned_data.get("import_mode") or "sync").strip()

        if applicants_file:
            result["applicants"] = _handle_imported_applicants(
                applicants_file,
                mode=import_mode,
            )
            messages.success(
                request,
                (
                    f"تم استيراد المتقدمين الجدد بنجاح "
                    f"(Batch #{result['applicants']['batch']}) — "
                    f"إضافة: {result['applicants']['created']} | "
                    f"تحديث: {result['applicants']['updated']} | "
                    f"تخطي: {result['applicants']['skipped']}"
                ),
            )
            if result["applicants"].get("portal_phase_switched"):
                messages.success(request, "تم تحويل البوابة تلقائيًا إلى مرحلة المتقدمين الجدد.")

        if schools_file:
            result["schools"] = _handle_imported_schools(
                schools_file,
                mode=import_mode,
            )
            messages.success(
                request,
                (
                    f"تم استيراد المدارس بنجاح "
                    f"(Batch #{result['schools']['batch']}) — "
                    f"إضافة: {result['schools']['created']} | "
                    f"تحديث: {result['schools']['updated']} | "
                    f"تخطي: {result['schools']['skipped']}"
                ),
            )

    return render(request, "portal/admin_import.html", {"form": form, "result": result})


@staff_member_required
@require_http_methods(["GET", "POST"])
def admin_import_new_applicants_view(request):
    if request.method == "GET":
        return redirect("portal:admin_import")

    uploaded_file = request.FILES.get("applicants_file")
    if not uploaded_file:
        messages.error(request, "فضلاً اختر ملف المتقدمين الجدد أولاً.")
        return redirect("portal:admin_import")

    import_mode = (request.POST.get("import_mode") or "sync").strip()
    result = _handle_imported_applicants(uploaded_file, mode=import_mode)

    messages.success(
        request,
        (
            f"تم استيراد المتقدمين الجدد بنجاح "
            f"(Batch #{result['batch']}) — "
            f"إضافة: {result['created']} | "
            f"تحديث: {result['updated']} | "
            f"تخطي: {result['skipped']}"
        ),
    )
    messages.success(request, "تم تحويل البوابة تلقائيًا إلى مرحلة المتقدمين الجدد.")
    return redirect("portal:admin_import")


@staff_member_required
@require_http_methods(["GET", "POST"])
def admin_import_schools_view(request):
    if request.method == "GET":
        return redirect("portal:admin_import")

    uploaded_file = request.FILES.get("schools_file")
    if not uploaded_file:
        messages.error(request, "فضلاً اختر ملف المدارس أولاً.")
        return redirect("portal:admin_import")

    import_mode = (request.POST.get("import_mode") or "sync").strip()
    result = _handle_imported_schools(uploaded_file, mode=import_mode)

    messages.success(
        request,
        (
            f"تم استيراد المدارس بنجاح "
            f"(Batch #{result['batch']}) — "
            f"إضافة: {result['created']} | "
            f"تحديث: {result['updated']} | "
            f"تخطي: {result['skipped']}"
        ),
    )
    return redirect("portal:admin_import")


@staff_member_required
def admin_new_applicants_sorting_view(request):
    base_qs = (
        Application.objects
        .select_related("applicant", "achieved_pref__vacancy")
        .prefetch_related("prefs", "prefs__vacancy")
        .filter(
            applicant__is_active=True,
            status="submitted",
        )
        .order_by("applicant__sector", "applicant__gender", "submitted_at", "id")
    )

    applications = [app for app in base_qs if not _is_official_proxy(app.applicant)]

    rows = []
    for app in applications:
        prefs = list(app.prefs.select_related("vacancy").order_by("rank", "id"))
        rows.append({
            "app": app,
            "applicant": app.applicant,
            "prefs": prefs[:5],
            "prefs_count": len(prefs),
            "achieved_pref": getattr(app, "achieved_pref", None),
        })

    ctx = {
        "rows": rows,
        "total_applications": len(applications),
        "total_open_vacancies": SchoolVacancy.objects.filter(
            is_open=True
        ).exclude(deputy_need=0).count(),
        "total_assigned": sum(1 for app in applications if getattr(app, "achieved_pref_id", None)),
        "total_unassigned": sum(1 for app in applications if not getattr(app, "achieved_pref_id", None)),
        "message": (
            "هذه شاشة فرز المتقدمين الجدد. "
            "إذا كان العدد صفرًا فهذا يعني أنه لا توجد طلبات مرسلة من المتقدمين الجدد حتى الآن."
        ),
    }
    return render(request, "portal/admin_new_applicants_sorting.html", ctx)


@staff_member_required
@require_POST
@transaction.atomic
def admin_run_new_applicants_sorting_view(request):
    result = _run_new_applicants_sorting(decided_by=request.user)
    messages.success(
        request,
        f"تم فرز المتقدمين الجدد بنجاح. إجمالي الطلبات: {result['applications']}، الموزع: {result['assigned']}، غير الموزع: {result['unassigned']}.",
    )
    return redirect("portal:admin_dashboard")


# =========================================================
# Admin: Non Applicants / Incomplete Applications
# =========================================================
def _admin_non_applicants_base_qs(q: str):
    qs = (
        Applicant.objects
        .filter(is_active=True)
        .select_related("application")
        .prefetch_related("application__prefs")
    )

    if q:
        qs = qs.filter(
            Q(full_name__icontains=q)
            | Q(national_id__icontains=q)
            | Q(sector__icontains=q)
            | Q(mobile__icontains=q)
            | Q(current_school__icontains=q)
        )

    return qs


def _apply_non_applicants_mode(qs, mode: str):
    if mode == "none":
        return qs.filter(application__isnull=True)

    if mode == "entered_not_confirmed":
        return qs.filter(
            application__isnull=False,
            application__confirmed_at__isnull=True,
        )

    if mode == "confirmed_not_submitted":
        return qs.filter(
            application__confirmed_at__isnull=False,
        ).exclude(application__status="submitted")

    if mode == "submitted_without_prefs":
        return qs.filter(
            application__status="submitted",
            application__prefs__isnull=True,
        ).distinct()

    if mode == "submitted_with_prefs":
        return qs.filter(
            application__status="submitted",
            application__prefs__isnull=False,
        ).distinct()

    if mode == "locked_incomplete":
        return qs.filter(
            application__locked=True,
        ).exclude(application__status="submitted")

    # الافتراضي: كل من لا يملك إرسالًا نهائيًا
    return qs.filter(
        Q(application__isnull=True)
        | ~Q(application__status="submitted")
    ).distinct()


@staff_member_required
def admin_non_applicants_view(request):
    q = (request.GET.get("q") or "").strip()
    mode = (request.GET.get("mode") or "not_submitted").strip()

    base_qs = _admin_non_applicants_base_qs(q)
    qs = _apply_non_applicants_mode(base_qs, mode).order_by("-id")

    page_obj = _paginate(request, qs, per_page=50)

    for applicant in page_obj.object_list:
        app = getattr(applicant, "application", None)
        code = _application_progress_code(app)
        applicant.progress_code = code
        applicant.progress_label = _application_progress_label(code)
        applicant.progress_note = _application_progress_note(code)
        applicant.prefs_count = app.prefs.count() if app else 0
        applicant.application_obj = app

    total_active = Applicant.objects.filter(is_active=True).count()
    total_submitted = Applicant.objects.filter(
        is_active=True,
        application__status="submitted",
    ).distinct().count()
    total_not_submitted = Applicant.objects.filter(is_active=True).filter(
        Q(application__isnull=True) | ~Q(application__status="submitted")
    ).distinct().count()
    total_none = Applicant.objects.filter(is_active=True, application__isnull=True).count()
    total_entered_not_confirmed = Applicant.objects.filter(
        is_active=True,
        application__isnull=False,
        application__confirmed_at__isnull=True,
    ).distinct().count()
    total_confirmed_not_submitted = Applicant.objects.filter(
        is_active=True,
        application__confirmed_at__isnull=False,
    ).exclude(application__status="submitted").distinct().count()
    total_submitted_without_prefs = Applicant.objects.filter(
        is_active=True,
        application__status="submitted",
        application__prefs__isnull=True,
    ).distinct().count()
    total_submitted_with_prefs = Applicant.objects.filter(
        is_active=True,
        application__status="submitted",
        application__prefs__isnull=False,
    ).distinct().count()
    total_locked_incomplete = Applicant.objects.filter(
        is_active=True,
        application__locked=True,
    ).exclude(application__status="submitted").distinct().count()

    win = PortalWindow.get()
    portal_open_now, _portal_msg, _ = _portal_gate()

    current_query = request.GET.urlencode()
    query_suffix = f"?{current_query}" if current_query else ""

    ctx = {
        "rows": page_obj,
        "q": q,
        "mode": mode,
        "total": qs.count(),
        "portal_window": win,
        "portal_open_now": portal_open_now,
        "current_query": current_query,
        "url_non_applicants_csv": f'{redirect("portal:admin_non_applicants_csv").url}{query_suffix}',
        "kpi": {
            "active": total_active,
            "submitted": total_submitted,
            "not_submitted": total_not_submitted,
            "none": total_none,
            "entered_not_confirmed": total_entered_not_confirmed,
            "confirmed_not_submitted": total_confirmed_not_submitted,
            "submitted_without_prefs": total_submitted_without_prefs,
            "submitted_with_prefs": total_submitted_with_prefs,
            "locked_incomplete": total_locked_incomplete,
        },
        "mode_labels": {
            "not_submitted": "غير مكتمل / لم يرسل",
            "none": "لم يدخل البوابة",
            "entered_not_confirmed": "دخل ولم يؤكد",
            "confirmed_not_submitted": "أكد ولم يرسل",
            "submitted_without_prefs": "مرسل بلا رغبات",
            "submitted_with_prefs": "مرسل برغبات",
            "locked_incomplete": "مقفل إداريًا",
        },
    }
    return render(request, "portal/admin_non_applicants.html", ctx)


@staff_member_required
@require_POST
@transaction.atomic
def admin_lock_incomplete_applications_view(request):
    """
    تثبيت الإجراء بعد إغلاق فترة التقديم:
    - لا يحوّل غير المكتمل إلى مرسل.
    - لا ينسب للمتقدم اختيارًا أو إقرارًا لم يفعله.
    - يقفل الطلبات التي لها سجل Application ولم ترسل نهائيًا.
    """
    portal_open_now, _portal_msg, _ = _portal_gate()
    if portal_open_now:
        messages.error(request, "أغلق البوابة أولًا قبل تثبيت إقفال الطلبات غير المكتملة.")
        return redirect("portal:admin_non_applicants")

    qs = Application.objects.select_for_update().exclude(status="submitted")
    total = qs.count()
    updated = qs.filter(locked=False).update(locked=True)

    messages.success(
        request,
        f"تم تثبيت إقفال الطلبات غير المكتملة. إجمالي الطلبات غير المكتملة: {total}، وتم إقفال الجديد منها: {updated}."
    )
    return redirect("portal:admin_non_applicants")


@staff_member_required
def admin_non_applicants_csv_view(request):
    q = (request.GET.get("q") or "").strip()
    mode = (request.GET.get("mode") or "not_submitted").strip()

    qs = _apply_non_applicants_mode(_admin_non_applicants_base_qs(q), mode).order_by("-id")

    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = 'attachment; filename="non_applicants.csv"'
    resp.write("\ufeff")

    w = csv.writer(resp)
    w.writerow([
        "#",
        "الاسم",
        "السجل",
        "الجوال",
        "القطاع",
        "الجنس",
        "المدرسة الحالية",
        "التصنيف الإجرائي",
        "الملاحظة الإدارية",
        "حالة الطلب الخام",
        "مقفل إداريًا؟",
        "عدد الرغبات",
        "تاريخ تأكيد البيانات",
        "تاريخ الإرسال",
    ])

    for i, applicant in enumerate(qs, start=1):
        app = getattr(applicant, "application", None)
        code = _application_progress_code(app)
        prefs_count = app.prefs.count() if app else 0
        w.writerow([
            i,
            applicant.full_name,
            applicant.national_id,
            applicant.mobile,
            applicant.sector,
            applicant.gender,
            applicant.current_school,
            _application_progress_label(code),
            _application_progress_note(code),
            getattr(app, "status", "") if app else "",
            "نعم" if app and getattr(app, "locked", False) else "لا",
            prefs_count,
            _fmt_dt(getattr(app, "confirmed_at", None)) if app else "",
            _fmt_dt(getattr(app, "submitted_at", None)) if app else "",
        ])

    return resp
