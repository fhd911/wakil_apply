from __future__ import annotations

from dataclasses import dataclass
from typing import Tuple

from openpyxl import load_workbook

from .models import Applicant, SchoolVacancy, ImportBatch


# =========================
# Result Object
# =========================
@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0


# =========================
# Normalizers
# =========================
def norm_text(s: str) -> str:
    s = (s or "").strip()
    s = " ".join(s.split())
    return s


def norm_gender(g: str) -> str:
    g = norm_text(g)
    mapping = {
        "ذكور": "بنين",
        "ذكر": "بنين",
        "اولاد": "بنين",
        "أولاد": "بنين",
        "بنين": "بنين",
        "إناث": "بنات",
        "اناث": "بنات",
        "أناث": "بنات",
        "انثى": "بنات",
        "بنات": "بنات",
    }
    return mapping.get(g, g)


def _to_int(v) -> int:
    try:
        return int(v or 0)
    except Exception:
        return 0


def _norm_import_mode(mode: str) -> str:
    mode = norm_text(mode or "sync").lower()
    allowed = {"sync", "create_only", "update_only"}
    return mode if mode in allowed else "sync"


def _cell(row, idx, default=""):
    try:
        value = row[idx]
    except Exception:
        return default
    return value if value is not None else default


# =========================
# Import Applicants (A..I)
# modes:
# - sync        = إضافة الجديد + تحديث الموجود
# - create_only = إضافة الجديد فقط
# - update_only = تحديث الموجود فقط
# ملاحظة تشغيلية: عند تحديث شاغر موجود لا نعيد فتحه تلقائيًا.
# الشاغر الجديد فقط يبدأ مفتوحًا، أما الشاغر المغلق فيبقى مغلقًا حتى يُفتح يدويًا.
# =========================
def import_applicants_xlsx(path: str, mode: str = "sync") -> Tuple[ImportBatch, ImportResult]:
    mode = _norm_import_mode(mode)

    wb = load_workbook(path)
    ws = wb.active

    batch = ImportBatch.objects.create(kind="applicants", file_name=path)
    res = ImportResult()

    for row in ws.iter_rows(min_row=2, values_only=True):
        full_name = norm_text(_cell(row, 0, ""))
        national_id = norm_text(str(_cell(row, 1, "")))
        mobile = norm_text(str(_cell(row, 2, "")))
        gender = norm_gender(_cell(row, 3, ""))
        current_job = norm_text(_cell(row, 4, ""))
        sector = norm_text(_cell(row, 5, ""))
        rank = norm_text(_cell(row, 6, ""))
        start_date = norm_text(str(_cell(row, 7, "")))
        current_school = norm_text(_cell(row, 8, ""))

        if not national_id:
            res.skipped += 1
            continue

        existing = Applicant.objects.filter(national_id=national_id).first()

        if mode == "create_only" and existing:
            res.skipped += 1
            continue

        if mode == "update_only" and not existing:
            res.skipped += 1
            continue

        _, was_created = Applicant.objects.update_or_create(
            national_id=national_id,
            defaults=dict(
                full_name=full_name,
                mobile=mobile,
                gender=gender,
                current_job=current_job,
                sector=sector,
                rank=rank,
                start_date=start_date,
                current_school=current_school,
                batch=batch,
                is_active=True,
            ),
        )

        if was_created:
            res.created += 1
        else:
            res.updated += 1

    return batch, res


# =========================
# Import Schools/Vacancies (A..R)
# modes:
# - sync        = إضافة الجديد + تحديث الموجود
# - create_only = إضافة الجديد فقط
# - update_only = تحديث الموجود فقط
# =========================
def import_schools_xlsx(path: str, mode: str = "sync") -> Tuple[ImportBatch, ImportResult]:
    mode = _norm_import_mode(mode)

    wb = load_workbook(path)
    ws = wb.active

    batch = ImportBatch.objects.create(kind="schools", file_name=path)
    res = ImportResult()

    for row in ws.iter_rows(min_row=2, values_only=True):
        ministry_no = norm_text(str(_cell(row, 0, "")))
        school_name = norm_text(_cell(row, 1, ""))
        stage = norm_text(_cell(row, 2, ""))
        sector = norm_text(_cell(row, 3, ""))
        establishment_status = norm_text(_cell(row, 4, ""))
        gender = norm_gender(_cell(row, 5, ""))
        education_type = norm_text(_cell(row, 6, ""))
        manager_national_id = norm_text(str(_cell(row, 7, "")))
        manager_name = norm_text(_cell(row, 8, ""))

        students_total = _to_int(_cell(row, 9, 0))
        classes_total = _to_int(_cell(row, 10, 0))
        students_metric = _to_int(_cell(row, 11, 0))
        class_metric = _to_int(_cell(row, 12, 0))
        stage_code = norm_text(str(_cell(row, 13, "")))
        stage_metric = _to_int(_cell(row, 14, 0))

        deputy_staff = _to_int(_cell(row, 15, 0))
        deputy_existing = _to_int(_cell(row, 16, 0))
        deputy_need = _to_int(_cell(row, 17, 0))

        if not school_name:
            res.skipped += 1
            continue

        key = ministry_no or school_name
        existing = SchoolVacancy.objects.filter(ministry_no=key).first()

        if mode == "create_only" and existing:
            res.skipped += 1
            continue

        if mode == "update_only" and not existing:
            res.skipped += 1
            continue

        obj, was_created = SchoolVacancy.objects.update_or_create(
            ministry_no=key,
            defaults=dict(
                ministry_no=key,
                school_name=school_name,
                stage=stage,
                sector=sector,
                establishment_status=establishment_status,
                gender=gender,
                education_type=education_type,
                manager_national_id=manager_national_id,
                manager_name=manager_name,
                students_total=students_total,
                classes_total=classes_total,
                students_metric=students_metric,
                class_metric=class_metric,
                stage_code=stage_code,
                stage_metric=stage_metric,
                deputy_staff=deputy_staff,
                deputy_existing=deputy_existing,
                deputy_need=deputy_need,
                # لا نضع is_open هنا؛ حتى لا يفتح الاستيراد شاغرًا أُغلق بسبب ترشيح سابق.
                batch=batch,
            ),
        )

        if was_created:
            # الشاغر الجديد يبدأ مفتوحًا بشكل طبيعي.
            if not obj.is_open:
                obj.is_open = True
                obj.save(update_fields=["is_open"])
            res.created += 1
        else:
            # الشاغر الموجود يحافظ على حالته الحالية: مفتوح أو مغلق.
            res.updated += 1

    return batch, res